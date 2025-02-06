import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from glob import glob
from pymorphy2 import MorphAnalyzer
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import NamedStyle
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore', category=FutureWarning)
import numpy as np
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import pandas as pd
import re
import hashlib

morph = MorphAnalyzer()

# Словарь месяцев
months = {
    'январь': 1,
    'февраль': 2,
    'март': 3,
    'апрель': 4,
    'май': 5,
    'июнь': 6,
    'июль': 7,
    'август': 8,
    'сентябрь': 9,
    'октябрь': 10,
    'ноябрь': 11,
    'декабрь': 12
}

def generate_hash(s, length=2):
    """
    Генерирует хэш для строки s и возвращает первые length символов.
    """
    if not isinstance(s, str):
        s = str(s)
    h = hashlib.sha256(s.encode()).hexdigest()
    return h[:length]


def extract_month_from_filename(filename):
    # Получаем базовое имя файла без расширения
    basename = os.path.basename(filename)
    basename_no_ext = os.path.splitext(basename)[0]

    # Приводим к нижнему регистру
    basename_no_ext = basename_no_ext.lower()

    # Убираем знаки препинания и заменяем символы подчеркивания на пробелы
    basename_no_ext = re.sub(r'[^\w\s]', ' ', basename_no_ext)
    basename_no_ext = basename_no_ext.replace('_', ' ')

    # Разбиваем на слова
    words = re.findall(r'\w+', basename_no_ext)

    # Словарь месяцев
    # months = {
    #     "январь": 1, "февраль": 2, "март": 3, "апрель": 4, "май": 5,
    #     "июнь": 6, "июль": 7, "август": 8, "сентябрь": 9,
    #     "октябрь": 10, "ноябрь": 11, "декабрь": 12
    # }

    for word in words:
        # Получаем нормальную форму слова
        parsed = morph.parse(word)[0]
        normal_form = parsed.normal_form
        if normal_form in months:
            return months[normal_form], normal_form

    # Если месяц не найден
    return None, None


def excel_column_to_index(column_letter):
    """
    Converts an Excel column letter to a zero-based index.

    Args:
        column_letter (str): The Excel column letter (e.g., 'A', 'B', ..., 'Z', 'AA', ...).

    Returns:
        int: The zero-based column index.
    """
    column_letter = column_letter.upper()  # Ensure uppercase for consistency
    index = 0
    for char in column_letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1  # Convert to zero-based index

def get_yellow_columns(file_path, sheet_name):
    # Загружаем книгу и выбираем лист
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]

    yellow_columns = []

    # Проход по всем колонкам и строкам листа
    for column in sheet.iter_cols():
        yellow_count = 0
        for cell in column:
            # Пропускаем объединенные ячейки
            if isinstance(cell, openpyxl.cell.cell.MergedCell) or sheet.row_dimensions[cell.row].hidden:
                # continue
                continue
            # Проверяем, есть ли заливка у ячейки
            if cell.fill:
                fill_color = cell.fill.fgColor
                # print(f"{cell.column_letter}-{cell.coordinate} {fill_color.rgb}")
                # Проверка RGB значения на желтый цвет, если тип цвета 'rgb'
                if fill_color and fill_color.type == "rgb" and fill_color.rgb == "FFFFFF00":
                    yellow_count += 1

        # Если в колонке более 15 строк окрашены в желтый
        if yellow_count > 9:
            column_number = column[0].column
            yellow_columns.append(column_number-1)
            continue

    # Выводим список букв колонок, в которых более 5 ячеек с желтой заливкой
    if yellow_columns:
        #add horizonts to panel
        yellow_columns.insert(0, yellow_columns[:1][0] - 1)
        yellow_column_letters = [get_column_letter(col+1) for col in yellow_columns]
        print(f"Колонки с большиством ячеек с желтой заливкой: {', '.join(yellow_column_letters)}")
    else:
        print("Нет колонок с более чем 15 ячейками с желтой заливкой")

    # Возвращаем список номеров колонок с желтой заливкой
    return yellow_columns


# Функция для проверки строки
def is_valid_row(row):
    # Указание индексов столбцов для проверки
    columns_to_check = ['Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg']  # Или 'C', 'D', 'R', 'F', 'G', 'H'
    # columns_to_check = ['Ruda', 'Cu', 'fRuda', 'fCu', 'fAg']  # Или 'C', 'D', 'R', 'F', 'G', 'H'

    # Проверяем только указанные столбцы
    for col in columns_to_check:
        value = row[col]
        if isinstance(value, int) and value != 0:
            continue
        return True  # Если хотя бы одно значение не целое или равно 0
    return False  # Если все значения являются целыми и не равны 0


def fill_merged_cells_in_first_yellow_column1(sheet, first_column):
    """
    Заполняет объединённые ячейки в указанной колонке.

    :param sheet: объект листа Excel.
    :param first_column: индекс первой жёлтой колонки (1-based).
    """
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_col == first_column and merged_range.max_col == first_column:  # Только указанная колонка
            start_row = merged_range.min_row
            end_row = merged_range.max_row
            value = sheet.cell(row=start_row, column=first_column).value  # Читаем значение верхней левой ячейки

            # Разъединяем объединённые ячейки
            sheet.unmerge_cells(str(merged_range))
            for row in range(start_row, end_row + 1):
                if not sheet.row_dimensions[row].hidden:  # Пропускаем скрытые строки
                    sheet.cell(row=row, column=first_column).value = value  # Устанавливаем значение



def fill_merged_cells_in_first_yellow_column(sheet, first_column):
    """
    Заполняет объединённые ячейки в указанной колонке.

    :param sheet: объект листа Excel.
    :param first_column: индекс первой жёлтой колонки (1-based).
    """
    # Собираем список объединённых диапазонов в указанном столбце
    ranges_to_unmerge = [
        merged_range
        for merged_range in sheet.merged_cells.ranges
        if merged_range.min_col == first_column and merged_range.max_col == first_column
    ]

    # Обрабатываем каждый диапазон отдельно
    for merged_range in ranges_to_unmerge:
        start_row = merged_range.min_row
        end_row = merged_range.max_row
        value = sheet.cell(row=start_row, column=first_column).value  # Читаем значение верхней левой ячейки

        # Пытаемся разъединить ячейки, игнорируя ошибки KeyError
        try:
            sheet.unmerge_cells(str(merged_range))
        except KeyError:
            pass

        # Заполняем ранее объединённые ячейки значением из верхней левой
        for row in range(start_row, end_row + 1):
            if not sheet.row_dimensions[row].hidden:  # Пропускаем скрытые строки
                sheet.cell(row=row, column=first_column).value = value


# from openpyxl import load_workbook


def unmerge_horizontal_cells(sheet):
    # # Load the workbook and select the sheet
    # workbook = load_workbook(file_path)
    # sheet = workbook[sheet_name]

    # Iterate over all merged cell ranges in the sheet
    for merged_cell_range in list(sheet.merged_cells.ranges):
        # Check if the merge is horizontal (spans multiple columns but only one row)
        if merged_cell_range.min_row == merged_cell_range.max_row:
            # Get the value from the first cell in the merged range
            first_cell = sheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col)
            value = first_cell.value

            # Check if the value contains the search string
            if value and "того" in str(value):
                # Unmerge the cells
                sheet.unmerge_cells(str(merged_cell_range))

                # Write the value from the first cell into all the cells that were previously merged
                for col in range(merged_cell_range.min_col, merged_cell_range.max_col + 1):
                    cell = sheet.cell(row=merged_cell_range.min_row, column=col)
                    cell.value = value






def convert_xls_to_xlsx(xls_file_path, xlsx_file_path):
    import pandas as pd
    # Читаем файл .xls с помощью xlrd
    df_dict = pd.read_excel(xls_file_path, sheet_name=None, engine='xlrd')
    # Пишем в файл .xlsx с помощью openpyxl
    with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def convert_xlsb_to_xlsx_with_formatting(xlsb_file_path, xlsx_file_path):
    """
    Конвертирует файл .xlsb в .xlsx с сохранением форматирования.
    Требуется установленный Microsoft Excel.
    """
    import win32com.client as win32
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        # Открываем файл .xlsb
        workbook = excel.Workbooks.Open(xlsb_file_path)
        # Сохраняем как .xlsx
        workbook.SaveAs(xlsx_file_path, FileFormat=51)  # 51 - это код для .xlsx
        workbook.Close()
    except Exception as e:
        print(f"Ошибка при конвертации файла {xlsb_file_path}: {e}")
    finally:
        excel.Quit()


def convert_xls_to_xlsx_with_formatting(xls_file_path, xlsx_file_path):
    """
    Конвертирует файл .xls в .xlsx с сохранением форматирования.
    Требуется установленный Microsoft Excel.
    """
    import win32com.client as win32
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(xls_file_path)
        workbook.SaveAs(xlsx_file_path, FileFormat=51)  # 51 - это формат .xlsx
        workbook.Close()
    except Exception as e:
        print(f"Ошибка при конвертации файла {xls_file_path}: {e}")
    finally:
        excel.Quit()


def filter_columns_with_whole_number_sums(df, columns_to_check=["Ruda",	"Cu",	"Ag",	"fRuda",	"fCu",	"fAg"]):
    """
    Filters out rows from a DataFrame where the sum of specified columns is an integer.

    Parameters:
        df (pd.DataFrame): The input DataFrame.
        columns_to_check (list): List of column names to check.

    Returns:
        pd.DataFrame: A DataFrame with rows removed where the sum of specified columns is an integer.
    """
    # Create a copy of the DataFrame to avoid modifying the original
    filtered_df = df.copy()

    # Iterate through each row
    for index, row in filtered_df.iterrows():
        # Calculate the sum of the specified columns for the current row
        row_sum = 0
        has_non_numeric = False

        for col in columns_to_check:
            value = row[col]

            # Skip if the cell is missing or empty
            if pd.isna(value) or value == '':
                continue

            # Check if the value is numeric
            try:
                float_value = float(value)  # Convert to float
                row_sum += float_value
            except (ValueError, TypeError):
                # If conversion to float fails, it's a non-numeric value
                has_non_numeric = True
                break  # No need to continue for this row

        # Skip rows with non-numeric values
        if has_non_numeric:
            # print(f"Row {index} contains non-numeric values and will be skipped.")
            continue

        # Check if the sum is an integer
        if row_sum == int(row_sum):  # Compare the sum to its integer version
            # print(f"Row {index} has an integer sum ({row_sum}) and will be removed.")
            filtered_df.drop(index, inplace=True)  # Drop the row

    return filtered_df


def forward_fill_column_by_index(sheet, column_index):
    """
    Forward fills empty cells in a specific column of a Worksheet or DataFrame using the column index.

    Parameters:
        sheet (Worksheet or pd.DataFrame): The input Worksheet or DataFrame.
        column_index (int): The index of the column to process.

    Returns:
        pd.DataFrame: The DataFrame with empty cells in the specified column forward-filled.
    """
    # Convert the Worksheet to a DataFrame if it's not already one
    if not isinstance(sheet, pd.DataFrame):
        # Read the Worksheet into a DataFrame
        data = sheet.values
        columns = next(data)  # First row is the header
        sheet = pd.DataFrame(data, columns=columns)

    # Create a copy of the DataFrame to avoid modifying the original
    filled_sheet = sheet.copy()

    # Iterate through the rows
    for i in range(len(filled_sheet) - 1):  # Stop at the second-to-last row
        current_value = filled_sheet.iat[i, column_index]
        next_value = filled_sheet.iat[i + 1, column_index]

        # Check if the next value is empty (NaN or empty string)
        if pd.isna(next_value) or next_value == '':
            # Fill the next row with the current value
            filled_sheet.iat[i + 1, column_index] = current_value

    return filled_sheet



def load_excel_data_with_flex(folder_path,tip=1):
    """Загружает данные из Excel файлов, учитывая возможные вариации в названиях месяцев в именах файлов."""
    # Шаг 1: Конвертируем все .xls файлы в .xlsx
    xls_files = glob(os.path.join(folder_path, "*.xls"))
    for xls_file in xls_files:
        # Генерируем путь для .xlsx файла
        xlsx_file = os.path.splitext(xls_file)[0] + '.xlsx'
        # Проверяем, существует ли уже конвертированный файл, чтобы избежать повторной конвертации
        if not os.path.exists(xlsx_file):
            try:
                convert_xls_to_xlsx_with_formatting(xls_file, xlsx_file)
                print(f"Конвертирован файл {xls_file} в {xlsx_file}")
            except Exception as e:
                print(f"Ошибка при конвертации файла {xls_file}: {e}")
                continue  # Переходим к следующему файлу, если конвертация не удалась

    xlsb_files = glob(os.path.join(folder_path, "*.xlsb"))
    for xlsb_file in xlsb_files:
        # Генерируем путь для .xlsx файла
        xlsx_file = os.path.splitext(xlsb_file)[0] + '.xlsx'
        # Проверяем, существует ли уже конвертированный файл, чтобы избежать повторной конвертации
        if not os.path.exists(xlsx_file):
            try:
                convert_xlsb_to_xlsx_with_formatting(xlsb_file, xlsx_file)
                print(f"Конвертирован файл {xlsb_file} в {xlsx_file}")
            except Exception as e:
                print(f"Ошибка при конвертации файла {xlsb_file}: {e}")
                continue  # Переходим к следующему файлу, если конвертация не удалась

    # Шаг 2: Обрабатываем только .xlsx файлы
    xlsx_files = glob(os.path.join(folder_path, "*.xlsx"))

    final_columns = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg', 'Uchastok', 'month', 'Block']
    final_df = pd.DataFrame(columns=final_columns)

    for file_path in xlsx_files:
        file_name = os.path.basename(file_path)
        month_num, month_name = extract_month_from_filename(file_name)
        if not month_num:
            print(f"-> -> ->  нет месяца - пропускаю файл: {file_name}")
            continue  # Skip files without recognizable month

        normalized_path = os.path.normpath(folder_path)
        # Извлекаем имя последней папки
        last_folder_name = os.path.basename(normalized_path)

        wb = load_workbook(file_path, data_only=True, read_only=False)
        # sheet_names = wb.sheetnames[1:]  # Exclude the first sheet

        # Список всех видимых листов
        visible_sheets = [sheet_name for sheet_name in wb.sheetnames if wb[sheet_name].sheet_state == "visible"]

        if len(visible_sheets) > 1:
            # Пропуск первого видимого листа
            visible_sheets = visible_sheets[1:] if len(visible_sheets) > 1 else []

        for sheet_name in visible_sheets:
            ws = wb[sheet_name]

            print(f"\n>>>>>>>>>>>>>>>>   {last_folder_name} == {month_name} ===  {sheet_name}  <<<<<<<<<<<<<<<<<<<<<<<<<<")
            # print(file_name)

            # Skip empty sheets
            if ws.max_row == 0 or ws.max_column == 0:
                print(f"Sheet '{sheet_name}' is empty.")
                continue

            # Get yellow columns using the provided function
            yellow_cols = get_yellow_columns(file_path, sheet_name)
            if not yellow_cols:
                print(f"No yellow columns found in sheet '{sheet_name}' in file '{file_name}'.")
                continue

            # Берём индекс первой желтой колонки (0-based)
            first_yellow_index = yellow_cols[0]
            # Вычисляем индекс колонки непосредственно перед первой желтой
            extra_col_index = first_yellow_index - 1

            # Заполняем объединённые ячейки только в первой жёлтой колонке =+1 так как нумерация в ексель с 1
            fill_merged_cells_in_first_yellow_column(ws, first_yellow_index)
            fill_merged_cells_in_first_yellow_column(ws, first_yellow_index + 1)
            # и во второй
            fill_merged_cells_in_first_yellow_column(ws, first_yellow_index + 2)

            # unmerge and fill the "ИТОГО" for filtering
            unmerge_horizontal_cells(ws)

            # forward fill panel columns
            forward_fill_column_by_index(ws, first_yellow_index)

            # Try to load data and set columns
            data_rows = []
            for idx, row in enumerate(ws.iter_rows(values_only=False), start=ws.min_row):
                if ws.row_dimensions[idx].hidden:
                    # Пропускаем скрытые строки
                    continue
                # Извлекаем значения ячеек
                cell_values = [cell.value for cell in row]
                data_rows.append(cell_values)

            if not data_rows:
                print(f"Sheet '{sheet_name}' has no data after removing hidden rows.")
                continue

            # Replace None in header with empty string
            header = [cell if cell is not None else '' for cell in data_rows[0]]
            data = pd.DataFrame(data_rows[1:], columns=header)

            # Filter yellow columns
            data = data.iloc[:, yellow_cols]

            # Take the very first column returned by get_yellow_columns
            col_before_yellow = yellow_cols[0]

            # Debugging: Print the column name and type
            # print(f"Column before first yellow column: {col_before_yellow}")
            # print(f"Type of data[col_before_yellow]: {type(data.iloc[:, col_before_yellow])}")

            # # Create a 'Block' column by cumulatively counting 'Итог' occurrences
            # data['Block'] = (data.iloc[:, col_before_yellow].astype(str).str.lower().str.contains('того')).cumsum() + 1
            #
            # # Remove rows where 'Итог' is present in the column before the first yellow column
            # data = data[~data.iloc[:, col_before_yellow].astype(str).str.lower().str.contains('того')]

            # Normalize and debug

            # print("Raw column values:", data.iloc[:, col_before_yellow].unique())

            col_before_yellow = yellow_cols[0]-1

            matches = (
                data.iloc[:, col_before_yellow]
                .astype(str)  # Ensure all values are strings
                .str.strip()  # Remove leading/trailing spaces
                .str.lower()  # Normalize case
                .str.contains('того')  # Search for 'того'
            )

            # print("Matching rows:", matches)

            # Add Block column
            data['Block'] = matches.cumsum() + 1

            # Remove rows with 'того'
            data = data[~matches]

            # Set appropriate columns based on the sheet name
            if 'ОГР' in sheet_name.upper():
                ogr_columns = ['Panel', 'Ruda', 'Cu', 'fRuda', 'fCu', 'Ag', 'fAg']
                if len(data.columns) == len(ogr_columns):
                    data.columns = ogr_columns
                    data['Shtrek'] = ''  # Add empty 'Shtrek' column
                else:
                    print(f"Warning: Sheet '{sheet_name}' has unexpected number of columns for OGR sheet.")
                    continue
            else:
                match tip:
                    case 1:
                        standard_columns = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg', 'Block']   # STANDART
                    case 2:
                        standard_columns = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'fRuda', 'fCu', 'Ag', 'fAg', 'Block']   # Nurkazgan
                if len(data.columns) == len(standard_columns):
                    data.columns = standard_columns
                else:
                    print(f"Warning: Sheet '{sheet_name}' has unexpected number of columns.")
                    continue

            # Fill down the 'Panel' values
            # data['Panel'].ffill(inplace=True)

            # Filter out rows where 'Shtrek' is empty (only if not OGR sheet)
            if 'ОГР' not in sheet_name.upper():
                data = data[data['Shtrek'].notnull() & (data['Shtrek'] != '')]

            # if 'ОГР' in sheet_name.upper():
            # убираем строки с пустой панелью
            data = data[data['Panel'].notnull() & (data['Panel'] != '')]

            # Удаление первой строки
            data = data.iloc[1:]  # Пропускаем первую строку

            # Применение фильтра только целые числа
            # data = data[data.apply(is_valid_row, axis=1)]

            #filterout complete integers its not wo we need
            data = filter_columns_with_whole_number_sums(data)


            data['Uchastok'] = sheet_name
            data['month'] = month_name

            # Ensure all columns are in data
            for col in final_columns:
                if col not in data.columns:
                    data[col] = ''

            data = data[final_columns]
            # Filter rows where both Ruda and fRuda are 0
            data = data[~((data['Ruda'] == 0) & (data['fRuda'] == 0))]
            # data = data[(pd.notna(data['Ruda']) & pd.notna(data['fRuda'])) & ~((data['Ruda'] == 0) & (data['fRuda'] == 0))]

            # data = data[
            #     ~((data['Cu'].apply(lambda x: isinstance(x, int))) &
            #     (data['Ag'].apply(lambda x: isinstance(x, int))) &
            #     (data['fCu'].apply(lambda x: isinstance(x, int))) &
            #     (data['fAg'].apply(lambda x: isinstance(x, int)))
            #       )
            #     ]
            data = data[~(data['fAg'] == "кг") ]

            # String to search for
            search_string = 'того'

            # # Specific columns to check for the string
            # columns_to_check = ['Panel', 'Shtrek', 'Horizont']
            # # Filter out rows where the search string appears in any of the specified columns
            # data = data[~data[columns_to_check].apply(
            #     lambda row: row.astype(str).str.contains(search_string, case=False, na=False).any(), axis=1)]

            condition0 = data['Panel'].str.contains(search_string, case=False, na=False)
            condition1 = data['Shtrek'].str.contains(search_string, case=False, na=False)
            condition2 = data['Horizont'].str.contains(search_string, case=False, na=False)
            condition = condition0 | condition1 | condition2
            data = data[~condition]


            final_df = pd.concat([final_df, data], ignore_index=True)

            # final_df = add_calculated_columns(final_df)

    final_df = final_df[~((final_df['Ruda'] == 0) & (final_df['fRuda'] == 0))]

    # Add a new column for month order based on the mapping
    final_df['month_N'] = final_df['month'].map(months)
    # Sort the DataFrame by 'month_order' and then by 'Block'
    final_df = final_df.sort_values(by=['month_N', 'Block'])
    # Drop the temporary 'month_order' column if not needed
    final_df = final_df.drop(columns=['month_N'])

    # Save the final DataFrame to a single Excel file
    # Use folder name as output file name
    # t = datetime.now().microsecond
    # output_file_name = os.path.basename("_report "+folder_path.strip('/\\')) +str(t)+ '.xlsx'
    # output_file = os.path.join(folder_path, output_file_name)

    generate_report_with_charts(folder_path, final_df)
    return final_df



def add_summary_and_formula_rows1(output_file):
    # Load the workbook and select the first sheet
    workbook = load_workbook(output_file)
    sheet = workbook["По штрекам"]  # Adjust this to your first sheet name if different

    # Find the last row with data
    last_row = sheet.max_row

    # Define the columns of interest
    columns = {'p>0': 'Q', 'f>0': 'R', 'p=0': 'S', 'f=0': 'T'}

    # Add the sum row
    sum_row = last_row + 1
    for col_name, col_letter in columns.items():
        formula = f"=SUM({col_letter}2:{col_letter}{last_row})"
        sheet[f"{col_letter}{sum_row}"] = formula

    # Add the count row with the formula
    count_row = last_row + 2
    for col_name, col_letter in columns.items():
        if col_name in ['p>0', 'f>0']:  # Columns with ">0" condition
            formula = f'=COUNTIF({col_letter}2:{col_letter}{last_row},">0")'
            sheet[f"{col_letter}{count_row}"] = formula
        elif col_name in ['p=0', 'f=0']:  # Columns with "=0" condition
            formula = f'=COUNTIF({col_letter}2:{col_letter}{last_row},"=0")'
            sheet[f"{col_letter}{count_row}"] = formula


    # Add percentage calculations to the left of the sums row
    perc_column = 'U'  # Insert percentages in column P
    sheet[f"{perc_column}{sum_row}"] = f"=R{sum_row}/Q{sum_row}"  # R361/Q361
    sheet[f"{perc_column}{count_row}"] = f"=R{count_row}/Q{count_row}"  # R362/Q362

    # Format percentage cells
    percent_style = NamedStyle(name="percent_style", number_format="0.00%")
    workbook.add_named_style(percent_style)  # Ensure the style is added once
    sheet[f"{perc_column}{sum_row}"].style = "percent_style"
    sheet[f"{perc_column}{count_row}"].style = "percent_style"

    # Force recalculation in Excel
    workbook.properties.calcMode = "auto"

    # Save the workbook
    workbook.save(output_file)



def add_summary_and_formula_rows(output_file):
    # Load the workbook and select the first sheet
    workbook = load_workbook(output_file)
    sheet = workbook["По штрекам"]  # Adjust this to your first sheet name if different

    # Find the last row with data
    last_row = sheet.max_row

    # Define the columns for existing data
    columns = {'p': 'Q', 'f': 'R'}  # Assuming 'p' in column Q and 'f' in column R

    # Add titles for new columns
    sheet['S1'] = 'Bbln (p=1 and f=1)'
    sheet['T1'] = 'HeBbln (p=1 and f=0)'
    sheet['U1'] = 'BHE (p=0 and f=1)'

    # Add columns for PP, VN, NP
    pp_col, vn_col, np_col = 'S', 'U', 'T'  # Assign new columns for PP, VN, NP

    for row in range(2, last_row + 1):
        # PP = 1 when p=1 and f=1, otherwise 0
        sheet[f"{pp_col}{row}"] = f"=IF(AND({columns['p']}{row}=1,{columns['f']}{row}=1),1,0)"
        # VN = 1 when p=0 and f=1, otherwise 0
        sheet[f"{vn_col}{row}"] = f"=IF(AND({columns['p']}{row}=0,{columns['f']}{row}=1),1,0)"
        # NP = 1 when p=1 and f=0, otherwise 0
        sheet[f"{np_col}{row}"] = f"=IF(AND({columns['p']}{row}=1,{columns['f']}{row}=0),1,0)"

    # Add the sum row for new columns
    sum_row = last_row + 1
    for col in ["q","r",pp_col, vn_col, np_col]:
        formula = f"=SUM({col}2:{col}{last_row})"
        sheet[f"{col}{sum_row}"] = formula

    # Add the count row for new columns
    count_row = last_row + 2
    for col in [pp_col, vn_col, np_col]:
        formula = f'=COUNTIF({col}2:{col}{last_row}, ">0")'
        sheet[f"{col}{count_row}"] = formula

    # Add percentage calculations for new columns
    perc_column = 'V'  # Choose a column for percentage calculations

    # Ensure "percent_style" is added only once
    if "percent_style" not in workbook.named_styles:
        percent_style = NamedStyle(name="percent_style", number_format="0.00%")
        workbook.add_named_style(percent_style)


    sheet[f"{perc_column}{sum_row}"] = f"=SUM({vn_col}{sum_row},{np_col}{sum_row})/Q{sum_row}"
    sheet[f"{perc_column}{sum_row}"].style = "percent_style"
    # for i, col in enumerate([pp_col, vn_col, np_col], start=1):
    #     sheet[f"{perc_column}{sum_row}"] = f"=SUM({vn_col}{sum_row}:{vn_col}{sum_row + i})/{pp_col}{sum_row}"
    #
    #     # sheet[f"{perc_column}{sum_row + i}"] = f"={col}{sum_row}/Q{sum_row}"  # Percentage of each column with respect to p (column Q)
    #     # sheet[f"{perc_column}{count_row + i}"] = f"={col}{count_row}/Q{count_row}"
    #
    #     # Apply the percent style
    #     sheet[f"{perc_column}{sum_row + i}"].style = "percent_style"
    #     sheet[f"{perc_column}{count_row + i}"].style = "percent_style"

    # Force recalculation in Excel
    workbook.properties.calcMode = "auto"

    # Save the workbook
    workbook.save(output_file)




def generate_report_with_charts(folder_path, full_df):

    # # Удаление указанных столбцов из DataFrame перед записью в Excel
    # full_df = full_df.drop(columns=["diff-Ruda", "diff-Cu", "diff-Ag"], errors='ignore')
    # block_and_month_aggregated_df = block_and_month_aggregated_df.drop(columns=["Ruda-fRuda", "Cu-fCu", "Ag-fAg", "diff-Ruda", "diff-Cu", "diff-Ag"], errors='ignore')
    # block_aggregated_df = block_aggregated_df.drop(columns=["diff-Ruda", "diff-Cu", "diff-Ag", "%rel-Ruda", "%rel-Cu", "%rel-Ag"], errors='ignore')


    full_df = add_calculated_columns(full_df)


    # Aggregate data by Block
    block_aggregated_df = group_by_block(full_df)
    # block_aggregated_df = calculate_monthly_averages(block_aggregated_df)
    # block_aggregated_df = set_fields_values_formatting(block_aggregated_df)

    # Функции для агрегации и обработки данных
    block_and_month_aggregated_df = group_by_block_and_month(full_df)
    # block_and_month_aggregated_df = calculate_monthly_averages(block_and_month_aggregated_df)
    # block_and_month_aggregated_df = set_fields_values_formatting(block_and_month_aggregated_df)

    # По штрекам лист

    # full_df = calculate_monthly_averages(full_df)
    # full_df = set_fields_values_formatting(full_df)

    # Список столбцов, которые нужно перевести в процентный формат
    percent_cols = [
        "1-(Ruda-fRuda)/Ruda",
        "1-(Cu-fCu)/Cu",
        "1-(Ag-fAg)/Ag",
        "1-(%Cu-%fCu)/%Cu",
        "Товарная руда ",
        "Cu в руде",
        "Ag в руде",
        "% Cu"
    ]




    # monthly_percent_df = calculate_monthly_average_percentages(final_df)


    # Генерация имени выходного файла
    t = datetime.now().microsecond
    output_file_name = os.path.basename("_report " + folder_path.strip('/\\')) + str(t) + '.xlsx'
    output_file = os.path.join(folder_path, output_file_name)

    # Создаем Excel с данными и графиками
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Записываем таблицы в Excel
        full_df.to_excel(writer, sheet_name="По штрекам", index=False)
        block_and_month_aggregated_df.to_excel(writer, sheet_name="Сумм. по панелям", index=False)
        block_aggregated_df.to_excel(writer, sheet_name="Сумм. по блокам", index=False)  # New sheet


        # Вместо записи отдельных листов для "Групп. по панелям" и "Средн. по штрекам",
        # вызываем функцию для создания одного листа с графиками
        create_excel_with_charts_on_one_sheet(
            mean_values_groupped_by_month(block_and_month_aggregated_df),
            mean_values_groupped_by_month(full_df),
            block_aggregated_df[['Block', '1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']]
        , writer)


    workbook = load_workbook(output_file)
    # Создаем и регистрируем стиль процента, если он еще не добавлен
    if "percent_style" not in workbook.named_styles:
        percent_style = NamedStyle(name="percent_style", number_format="0.00%")
        workbook.add_named_style(percent_style)

    # Применяем процентный стиль к указанным столбцам на нужных листах
    for sheet_name in ["По штрекам", "Сумм. по панелям", "Сумм. по блокам"]:
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            header_row = 1
            header_to_col = {}
            # Поиск столбцов с нужными заголовками в первой строке
            for cell in ws[header_row]:
                if cell.value in percent_cols:
                    header_to_col[cell.value] = cell.column_letter
            # Применение процентного стиля к найденным столбцам, начиная со второй строки
            for col_letter in header_to_col.values():
                for row in range(header_row + 1, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    # Если значение числовое, применяем процентный стиль
                    if isinstance(cell.value, (int, float)):
                        cell.style = "percent_style"



    # Freeze cell C2

    sheet_name = "По штрекам"  # Specify the sheet where you want to freeze C2
    sheet = workbook[sheet_name]
    sheet.freeze_panes = "C2"  # Set the freeze panes to C2
    workbook.save(output_file)
    # Add summary and formula rows to the first sheet
    # add_summary_and_formula_rows(output_file)

    # return full_df






def create_excel_with_charts_on_one_sheet(df1, df2, df3, writer):
    """
    Adds charts and data for three DataFrames on one Excel sheet,
    placing charts to the right of the tables and ensuring offsets between elements.

    Args:
        df1 (pd.DataFrame): First DataFrame for the first chart.
        df2 (pd.DataFrame): Second DataFrame for the second chart.
        df3 (pd.DataFrame): Third DataFrame for the third chart (yearly sums by block).
        writer (pd.ExcelWriter): ExcelWriter instance for writing to the file.
    """
    # Format numbers in DataFrame to 2 decimal places
    df1 = df1.round(3)
    df2 = df2.round(3)
    df3 = df3.round(3)
    # df3 = df3[['Block', '1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']]

    # Load the workbook and create a sheet for charts
    wb = writer.book
    ws = wb.create_sheet("Графики")

    ### Adding Data for the First DataFrame ###
    # headers_df1 = list(df1.columns)
    # df1 = df1[['Месяц', '1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']]
    # headers_df2 = ['Месяц', '1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']
    headers_df1 = ["Месяц", "Товарная руда", "Cu в руде", "Ag в руде", "%Cu"]
    ws.append(headers_df1)
    headers_row_df1 = ws.max_row

    data_start_row_df1 = ws.max_row + 1
    for row in df1.itertuples(index=False, name=None):
        ws.append(row)
    data_end_row_df1 = ws.max_row
    data_max_col_df1 = len(df1.columns)

    chart1 = BarChart()
    chart1.title = "Средние отклонения сгруппированные по панелям в разрезе месяцев"
    chart1.y_axis.title = "Значения"
    chart1.x_axis.title = "Месяцы"
    chart1.height = 15
    chart1.width = 25
    chart1.dLbls = DataLabelList()
    chart1.dLbls.showVal = True

    data_range1 = Reference(ws, min_col=2, min_row=headers_row_df1, max_col=data_max_col_df1, max_row=data_end_row_df1)
    categories1 = Reference(ws, min_col=1, min_row=data_start_row_df1, max_row=data_end_row_df1)
    chart1.add_data(data_range1, titles_from_data=True)
    chart1.set_categories(categories1)

    chart_start_col1 = data_max_col_df1 + 3
    chart_start_row1 = headers_row_df1
    chart_start_cell1 = f"{get_column_letter(chart_start_col1)}{chart_start_row1}"
    ws.add_chart(chart1, chart_start_cell1)

    ws.append([])
    ws.append([])

    ### Adding Data for the Second DataFrame ###
    # headers_df2 = list(df2.columns)
    # df2 = df2[['Месяц', '1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']]
    # headers_df2 = ['Месяц', '1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']
    headers_df2 = ["Месяц", "Товарная руда", "Cu в руде", "Ag в руде", "%Cu"]

    ws.append(headers_df2)
    headers_row_df2 = ws.max_row

    data_start_row_df2 = ws.max_row + 1
    for row in df2.itertuples(index=False, name=None):
        ws.append(row)
    data_end_row_df2 = ws.max_row
    data_max_col_df2 = len(df2.columns)

    chart2 = BarChart()
    chart2.title = "Средние отклонения по всем штрекам в разрезе месяцев"
    chart2.y_axis.title = "Значения"
    chart2.x_axis.title = "Месяцы"
    chart2.height = 15
    chart2.width = 25
    chart2.dLbls = DataLabelList()
    chart2.dLbls.showVal = True

    data_range2 = Reference(ws, min_col=2, min_row=headers_row_df2, max_col=data_max_col_df2, max_row=data_end_row_df2)
    categories2 = Reference(ws, min_col=1, min_row=data_start_row_df2, max_row=data_end_row_df2)
    chart2.add_data(data_range2, titles_from_data=True)
    chart2.set_categories(categories2)

    chart_start_col2 = data_max_col_df2 + 3
    chart_start_row2 = headers_row_df2 + 20  # Расположение графика рядом со вторым блоком данных
    chart_start_cell2 = f"{get_column_letter(chart_start_col2)}{chart_start_row2}"
    ws.add_chart(chart2, chart_start_cell2)

    ws.append([])
    ws.append([])

    ### Adding Data for the Third DataFrame ###
    # headers_df3 = list(df3.columns)
    headers_df3 = ["Панель№","Товарная руда","Cu в руде","Ag в руде","%Cu"]
    ws.append(headers_df3)
    headers_row_df3 = ws.max_row

    data_start_row_df3 = ws.max_row + 1
    for row in df3.itertuples(index=False, name=None):
        ws.append(row)
    data_end_row_df3 = ws.max_row
    data_max_col_df3 = len(df3.columns)

    chart3 = BarChart()
    chart3.title = "Годовые суммы по блокам/панелям"
    chart3.y_axis.title = "Значения"
    chart3.x_axis.title = "Блок/Панель"
    chart3.height = 15
    chart3.width = 25
    chart3.dLbls = DataLabelList()
    chart3.dLbls.showVal = True

    data_range3 = Reference(ws, min_col=2, min_row=headers_row_df3, max_col=data_max_col_df3, max_row=data_end_row_df3)
    categories3 = Reference(ws, min_col=1, min_row=data_start_row_df3+1, max_row=data_end_row_df3)
    chart3.add_data(data_range3, titles_from_data=True)
    chart3.set_categories(categories3)

    chart_start_col3 = data_max_col_df3 + 3
    chart_start_row3 = data_end_row_df3 + 30
    chart_start_cell3 = f"{get_column_letter(chart_start_col3)}{chart_start_row3}"
    ws.add_chart(chart3, chart_start_cell3)




def add_calculated_columns(df):
    """
    Добавляет расчетные столбцы в DataFrame:
    - dRuda, dCu, dAg
    - %Ruda, %Cu, %Ag

    :param df: pandas DataFrame, содержащий исходные столбцы
    :return: pandas DataFrame с добавленными расчетными столбцами
    """
    df['Ruda'] = df['Ruda'].astype(str)  # Приводим к строковому типу
    df = df[~df['Ruda'].str.contains('Руда', na=False)]

    df['Ruda'] = df['Ruda'].fillna(0)
    df['Cu'] = df['Cu'].fillna(0)
    df['Ag'] = df['Ag'].fillna(0)
    df['fRuda'] = df['fRuda'].fillna(0)
    df['fCu'] = df['fCu'].fillna(0)
    df['fAg'] = df['fAg'].fillna(0)

    # Заменяем некорректные строки (например, 'None') на NaN
    df.replace('None', np.nan, inplace=True)


    # Преобразуем все указанные колонки в float
    columns_to_convert = ['Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg']

    for column in columns_to_convert:
        # Replace None and empty strings with NaN
        df[column] = df[column].replace([None, ''], np.nan)

        # Remove commas and strip whitespace
        df[column] = df[column].astype(str).str.replace(',', '').str.strip()

        # Convert to numeric, coercing errors to NaN
        df[column] = pd.to_numeric(df[column], errors='coerce')

        # Optional: Replace NaN with 0
        df[column] = df[column].fillna(0)

    # df[columns_to_convert] = df[columns_to_convert].astype(float)


    # Расчет dRuda, dCu, dAg (с обработкой ошибок, замена NaN на 0)
    df['diff-Ruda'] = abs(df['Ruda'] - df['fRuda'])
    df['diff-Cu'] = abs(df['Cu'] - df['fCu'])
    df['diff-Ag'] = abs(df['Ag'] - df['fAg'])
    # df['diff-Ruda'] = df['Ruda'] - df['fRuda']
    # df['diff-Cu'] = df['Cu'] - df['fCu']
    # df['diff-Ag'] = df['Ag'] - df['fAg']

    # Замена NaN на 0, если расчет невозможен
    df['diff-Ruda'] = df['diff-Ruda'].fillna(0)
    df['diff-Cu'] = df['diff-Cu'].fillna(0)
    df['diff-Ag'] = df['diff-Ag'].fillna(0)

    # Расчет %Ruda, %Cu, %Ag (с обработкой деления на ноль или NaN)
    # df['%rel-Ruda'] = abs(np.where(df['Ruda'] != 0, df['diff-Ruda'] / df['Ruda'], 1.0) * 100)  # 100% если Ruda == 0
    # df['%rel-Cu'] = abs(np.where(df['Cu'] != 0, df['diff-Cu'] / df['Cu'], 1.0) * 100)  # 100% если Cu == 0
    # df['%rel-Ag'] = abs(np.where(df['Ag'] != 0, df['diff-Ag'] / df['Ag'], 1.0) * 100)  # 100% если Ag == 0

    # Расчет %Ruda, %Cu, %Ag (с обработкой деления на ноль или NaN и ограничением максимум 100)
    df['1-(Ruda-fRuda)/Ruda'] = np.clip(np.where(df['Ruda'] != 0, 1 - df['diff-Ruda'] / df['Ruda'], 1.0), 0, 100)
    df['1-(Cu-fCu)/Cu'] = np.clip(np.where(df['Cu'] != 0, 1 - df['diff-Cu'] / df['Cu'], 1.0) , 0, 100)
    df['1-(Ag-fAg)/Ag'] = np.clip(np.where(df['Ag'] != 0, 1 - df['diff-Ag'] / df['Ag'], 1.0) , 0, 100)

    # Замена NaN на 100%, если расчет невозможен
    df['1-(Ruda-fRuda)/Ruda'] = df['1-(Ruda-fRuda)/Ruda'].fillna(1)
    df['1-(Cu-fCu)/Cu'] = df['1-(Cu-fCu)/Cu'].fillna(1)
    df['1-(Ag-fAg)/Ag'] = df['1-(Ag-fAg)/Ag'].fillna(1)

    # Добавляем новые расчёты: %Cu и %fCu
    df['%Cu'] = np.where(
        (df['Cu'] == 0) & (df['fCu'] != 0),
        0,
        np.clip(np.where(df['Ruda'] != 0, df['Cu'] / df['Ruda'], 1.0), 0, 100)
    )
    df['%fCu'] = np.where(
        (df['fCu'] == 0) & (df['Cu'] != 0),
        0,
        np.clip(np.where(df['fRuda'] != 0, df['fCu'] / df['fRuda'], 1.0), 0, 100)
    )
    # Расчёт аналогичных метрик для %Cu и %Ag
    df['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where((df['Ruda'] != 0) & (df['fRuda'] != 0),
                 (1 - ((df['Cu'] / df['Ruda'] - df['fCu'] / df['fRuda']) / ((df['Cu'] / df['Ruda'])) )),
                 1),
        0, 100)

    # Добавляем дополнительные колонки p>0, f>0, p=0, f=0
    df['p>0'] = np.where(df['Ruda'] > 0, 1, 0)
    df['f>0'] = np.where(df['fRuda'] > 0, 1, 0)
    # df['p=0'] = np.where(df['Ruda'] == 0, 1, 0)
    # df['f=0'] = np.where(df['fRuda'] == 0, 1, 0)

    return df


def set_fields_values_formatting(df):
    """
    Преобразует DataFrame monthly_avg:
    - Переименовывает колонки.
    - Конвертирует числовые колонки в числовой формат.
    - Сортирует строки по порядку месяцев.
    """
    # Обновлённое переименование колонок для 6 столбцов
    df.columns = ['Месяц', 'Товарная руда (СМТ)', 'Cu в руде', 'Ag в руде', 'содерж.Cu']

    # Преобразуем последние три колонки в числовой формат
    numeric_columns = ['Товарная руда (СМТ)', 'Cu в руде', 'Ag в руде', 'содерж.Cu']
    df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')

    # Создаём порядок месяцев
    month_order = [
        "январь", "февраль", "март", "апрель", "май", "июнь",
        "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"
    ]

    # Сортируем строки по порядку месяцев
    df['Месяц'] = pd.Categorical(df['Месяц'], categories=month_order, ordered=True)
    df = df.sort_values(by='Месяц')

    return df



def group_by_block_and_month(df):
    """
    Группирует DataFrame по месяцам и штрекам, суммирует указанные колонки
    и добавляет расчетные колонки.

    :param df: Исходный DataFrame
    :return: Новый DataFrame с расчетными колонками
    """
    # Указываем, какие колонки нужно суммировать
    sum_columns = ['Ruda', 'fRuda', 'Cu', 'fCu', 'Ag', 'fAg']

    # Группировка по месяцам и панелям, вычисление суммы
    # grouped = df.groupby(['month', 'Horizont','Panel'], as_index=False)[sum_columns].sum()
    grouped = df.groupby(['Block','month' ], as_index=False)[sum_columns].sum()


    # Добавляем расчетные колонки
    grouped['Ruda-fRuda'] = grouped['Ruda'] - grouped['fRuda']
    grouped['Cu-fCu'] = grouped['Cu'] - grouped['fCu']
    grouped['Ag-fAg'] = grouped['Ag'] - grouped['fAg']

    # Замена ошибок на 0
    grouped['Ruda-fRuda'] = grouped['Ruda-fRuda'].fillna(0)
    grouped['Cu-fCu'] = grouped['Cu-fCu'].fillna(0)
    grouped['Ag-fAg'] = grouped['Ag-fAg'].fillna(0)

    # Добавляем расчетные колонки в процентах с ограничением максимум 100
    grouped['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(grouped['Ruda'] != 0, 1-abs(grouped['Ruda-fRuda'] / grouped['Ruda']), 1.0), 0, 100
    )
    grouped['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(grouped['Cu'] != 0, 1-abs(grouped['Cu-fCu'] / grouped['Cu']), 1.0), 0, 100
    )
    grouped['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(grouped['Ag'] != 0, 1-abs(grouped['Ag-fAg'] / grouped['Ag']), 1.0), 0, 100
    )

    # Добавляем расчёт %Cu и %fCu для данной группировки
    grouped['%Cu'] = np.where(grouped['Ruda'] != 0, grouped['Cu'] / grouped['Ruda'] * 100, 0)
    grouped['%fCu'] = np.where(grouped['fRuda'] != 0, grouped['fCu'] / grouped['fRuda'] * 100, 0)

    grouped['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where((grouped['Ruda'] != 0) & (grouped['fRuda'] != 0),
                 (1 - ((grouped['Cu'] / grouped['Ruda'] - grouped['fCu'] / grouped['fRuda']) / ((grouped['Cu'] / grouped['Ruda'])) )),
                 1),
        0, 100)

    grouped = grouped[['Block','month', 'Ruda', 'fRuda', 'Cu', 'fCu', 'Ag', 'fAg','%Cu','%fCu','1-(Ruda-fRuda)/Ruda','1-(Cu-fCu)/Cu','1-(Ag-fAg)/Ag','1-(%Cu-%fCu)/%Cu']]
    grouped['p>0'] = np.where(grouped['Ruda'] > 0, 1, 0)
    grouped['f>0'] = np.where(grouped['fRuda'] > 0, 1, 0)

    return grouped



def mean_values_groupped_by_month(df):
    """
    Группирует DataFrame по месяцам, вычисляет абсолютное среднее значение для %tRuda, %tCu, %tAg
    и записывает вычитание из 1 в виде процентов в колонки Ruda, Cu, Ag.
    Исключает февраль и сентябрь из расчета.

    :param df: DataFrame, содержащий колонки month, %tRuda, %tCu, %tAg.
    :return: Новый DataFrame с агрегированными данными.
    """

    # # Средние значения по месяцам с расчётом абсолютных значений
    # monthly_avg = df.groupby('month', as_index=False)[['(Ruda-fRuda)/Ruda', '(Cu-fCu)/Cu', '(Ag-fAg)/Ag']].apply( lambda x: x.abs().mean() )
    monthly_avg = df.groupby('month', as_index=False)[
        ['1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']
    ].mean()
    # monthly_avg = df.groupby(['month', 'Block'], as_index=False)[['(Ruda-fRuda)/Ruda', '(Cu-fCu)/Cu', '(Ag-fAg)/Ag']].mean()


    # monthly_avg['aver-relation_Ruda'] = monthly_avg['1-(Ruda-fRuda)/Ruda']
    # monthly_avg['aver-relation_Cu'] = monthly_avg['1-(Cu-fCu)/Cu']
    # monthly_avg['aver-relation_Ag'] = monthly_avg['1-(Ag-fAg)/Ag']
    # monthly_avg['aver-%Cu'] = monthly_avg['%Cu']
    # monthly_avg['aver-%fCu'] = monthly_avg['%fCu']

    # monthly_avg['1-(%Cu-%fCu)/%Cu'] = np.clip(
    #     np.where((monthly_avg['Ruda'] != 0) & (monthly_avg['fRuda'] != 0),
    #              (1 - ((monthly_avg['Cu'] / monthly_avg['Ruda'] - monthly_avg['fCu'] / monthly_avg['fRuda']) / ((monthly_avg['Cu'] / monthly_avg['Ruda'])) )),
    #              1),
    #     0, 100)


    # Оставляем только нужные колонки
    # monthly_avg = monthly_avg[['month', 'aver-relation_Ruda', 'aver-relation_Cu', 'aver-relation_Ag', 'aver-%Cu', 'aver-%fCu']]
    monthly_avg = monthly_avg[['month', '1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']]
    return monthly_avg


def mean_values_groupped_by_month_and_block(df):
    """
    Группирует DataFrame по месяцам, вычисляет абсолютное среднее значение для %tRuda, %tCu, %tAg
    и записывает вычитание из 1 в виде процентов в колонки Ruda, Cu, Ag.
    Исключает февраль и сентябрь из расчета.

    :param df: DataFrame, содержащий колонки month, %tRuda, %tCu, %tAg.
    :return: Новый DataFrame с агрегированными данными.
    """

    # # Средние значения по месяцам с расчётом абсолютных значений
    # monthly_avg = df.groupby('month', as_index=False)[['(Ruda-fRuda)/Ruda', '(Cu-fCu)/Cu', '(Ag-fAg)/Ag']].apply( lambda x: x.abs().mean() )
    # monthly_avg = df.groupby('month', as_index=False)[
    #     ['1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']
    # ].mean()
    monthly_avg = df.groupby(['month', 'Block'], as_index=False)[['(Ruda-fRuda)/Ruda', '(Cu-fCu)/Cu', '(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']].mean()


    # monthly_avg['aver-relation_Ruda'] = monthly_avg['1-(Ruda-fRuda)/Ruda']
    # monthly_avg['aver-relation_Cu'] = monthly_avg['1-(Cu-fCu)/Cu']
    # monthly_avg['aver-relation_Ag'] = monthly_avg['1-(Ag-fAg)/Ag']
    # monthly_avg['aver-%Cu'] = monthly_avg['%Cu']
    # monthly_avg['aver-%fCu'] = monthly_avg['%fCu']

    # monthly_avg['1-(%Cu-%fCu)/%Cu'] = np.clip(
    #     np.where((monthly_avg['Ruda'] != 0) & (monthly_avg['fRuda'] != 0),
    #              (1 - ((monthly_avg['Cu'] / monthly_avg['Ruda'] - monthly_avg['fCu'] / monthly_avg['fRuda']) / ((monthly_avg['Cu'] / monthly_avg['Ruda'])) )),
    #              1),
    #     0, 100)


    # Оставляем только нужные колонки
    # monthly_avg = monthly_avg[['month', 'aver-relation_Ruda', 'aver-relation_Cu', 'aver-relation_Ag', 'aver-%Cu', 'aver-%fCu']]
    monthly_avg = monthly_avg[["block",'month', '1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']]

    # Создаём порядок месяцев
    month_order = [
        "январь", "февраль", "март", "апрель", "май", "июнь",
        "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"
    ]

    # Сортируем строки по порядку месяцев
    monthly_avg['Месяц'] = pd.Categorical(monthly_avg['Месяц'], categories=month_order, ordered=True)
    monthly_avg = monthly_avg.sort_values(by='Месяц')

    return monthly_avg



# def mean_values_groupped_by_month(df):
#     """
#     Группирует исходный DataFrame по месяцам, вычисляет средние значения для
#     %Ruda, %Cu, %Ag, а также для %Cu и %fCu,
#     и записывает эти значения в соответствующие колонки.
#     """
#     # Группировка по месяцам и вычисление средних значений
#     monthly_avg = df.groupby('month', as_index=False)[
#         ['1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '%Cu', '%fCu']
#     ].mean()
#
#     # Переименование и сохранение рассчитанных процентов в новые колонки
#     monthly_avg['otlonenie-Ruda'] = monthly_avg['1-(Ruda-fRuda)/Ruda']
#     monthly_avg['otlonenie-Cu'] = monthly_avg['1-(Cu-fCu)/Cu']
#     monthly_avg['otlonenie-Ag'] = monthly_avg['1-(Ag-fAg)/Ag']
#     monthly_avg['otlonenie-%Cu'] = monthly_avg['%Cu']
#     monthly_avg['otlonenie-%fCu'] = monthly_avg['%fCu']
#
#     # Оставляем только необходимые колонки в итоговом DataFrame
#     monthly_avg = monthly_avg[
#         ['month', 'otlonenie-Ruda', 'otlonenie-Cu', 'otlonenie-Ag', 'otlonenie-%Cu', 'otlonenie-%fCu']
#     ]
#
#
#     return monthly_avg



def group_by_block(df):
    """
     Группирует DataFrame по месяцам и штрекам, суммирует указанные колонки
     и добавляет расчетные колонки.

     :param df: Исходный DataFrame
     :return: Новый DataFrame с расчетными колонками
     """
    # Указываем, какие колонки нужно суммировать
    sum_columns = ['Ruda', 'fRuda', 'Cu', 'fCu', 'Ag', 'fAg']

    # Группировка по месяцам и панелям, вычисление суммы
    # grouped = df.groupby(['month', 'Horizont','Panel'], as_index=False)[sum_columns].sum()
    grouped = df.groupby(['Block', ], as_index=False)[sum_columns].sum()

    # Добавляем расчетные колонки
    grouped['Ruda-fRuda'] = grouped['Ruda'] - grouped['fRuda']
    grouped['Cu-fCu'] = grouped['Cu'] - grouped['fCu']
    grouped['Ag-fAg'] = grouped['Ag'] - grouped['fAg']

    # Замена ошибок на 0
    grouped['Ruda-fRuda'] = grouped['Ruda-fRuda'].fillna(0)
    grouped['Cu-fCu'] = grouped['Cu-fCu'].fillna(0)
    grouped['Ag-fAg'] = grouped['Ag-fAg'].fillna(0)

    # Добавляем расчёт %Cu и %fCu для данной группировки
    grouped['%Cu'] = np.where(grouped['Ruda'] != 0, grouped['Cu'] / grouped['Ruda'], 0)
    grouped['%fCu'] = np.where(grouped['fRuda'] != 0, grouped['fCu'] / grouped['fRuda'] , 0)


    # Добавляем расчетные колонки в процентах с ограничением максимум 100
    grouped['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(grouped['Ruda'] != 0, 1 - abs(grouped['Ruda-fRuda'] / grouped['Ruda']), 1.0), 0, 100
    )
    grouped['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(grouped['Cu'] != 0, 1 - abs(grouped['Cu-fCu'] / grouped['Cu']), 1.0), 0, 100
    )
    grouped['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(grouped['Ag'] != 0, 1 - abs(grouped['Ag-fAg'] / grouped['Ag']), 1.0), 0, 100
    )

    grouped['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where((grouped['Ruda'] != 0) & (grouped['fRuda'] != 0),
                 (1 - ((grouped['Cu'] / grouped['Ruda'] - grouped['fCu'] / grouped['fRuda']) / (
                 (grouped['Cu'] / grouped['Ruda'])))),
                 1),
        0, 100)

    grouped['p>0'] = np.where(grouped['Ruda'] > 0, 1, 0)
    grouped['f>0'] = np.where(grouped['fRuda'] > 0, 1, 0)

    # grouped = grouped[['Block', '1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']]
    #
    # grouped.columns= ['Панель', 'Товарная руда (СМТ)', 'Cu в руде', 'Ag в руде', 'содерж.Cu']

    return grouped



# Пример использования:
tip =1
# directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ЮЖР ИПГ 2024'
# directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Шатыркуль ИПГ 2024'
# directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ИПГ Саяк 3 2024'
# directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Жомарт ИПГ 2024'
# directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ИПГ Жайсан 2024'
directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ВЖР ИПГ 2024'
# directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Жиланды ИПГ 2024'
# directory =r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ЗР ИПГ 2024'
# directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Конырат ИПГ 2024'
# directory = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__Акбастау ИПГ 2024"
# directory = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__ИПГ 2024 С-1"
# directory,tip = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__Нурказган ИПГ 2024",2   # ----N
# directory,tip = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\Хаджиконган ИПГ 2024"   # ----N
# directory = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__Абыз ИПГ 2024"
final_data = load_excel_data_with_flex(directory,tip)