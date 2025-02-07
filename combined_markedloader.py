import os
import re
import hashlib
import warnings
from datetime import datetime
from glob import glob

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter

from pymorphy2 import MorphAnalyzer

warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=UserWarning)

try:
    import win32com.client as win32
except ImportError:
    win32 = None

# -----------------------
# Глобальные переменные
# -----------------------

morph = MorphAnalyzer()
months = {
    'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5, 'июнь': 6,
    'июль': 7, 'август': 8, 'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12
}
COL_LIMIT = 94


# -----------------------
# Функции конвертации
# -----------------------

def convert_xls_to_xlsx_with_formatting(xls_file_path, xlsx_file_path):
    """
    Конвертирует .xls в .xlsx, сохраняя форматирование (требует Windows + Excel).
    """
    if win32 is None:
        print("win32com недоступен; не могу конвертировать .xls.")
        return
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(xls_file_path)
        workbook.SaveAs(xlsx_file_path, FileFormat=51)  # 51 -> .xlsx
        workbook.Close()
    except Exception as e:
        print(f"Ошибка при конвертации {xls_file_path}: {e}")
    finally:
        excel.Quit()


def convert_xlsb_to_xlsx_with_formatting(xlsb_file_path, xlsx_file_path):
    """
    Конвертирует .xlsb в .xlsx, сохраняя форматирование (требует Windows + Excel).
    """
    if win32 is None:
        print("win32com недоступен; не могу конвертировать .xlsb.")
        return
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(xlsb_file_path)
        workbook.SaveAs(xlsx_file_path, FileFormat=51)  # 51 -> .xlsx
        workbook.Close()
    except Exception as e:
        print(f"Ошибка при конвертации {xlsb_file_path}: {e}")
    finally:
        excel.Quit()


# -----------------------
# Вспомогательные функции
# -----------------------

def generate_hash(s, length=2):
    if not isinstance(s, str):
        s = str(s)
    h = hashlib.sha256(s.encode()).hexdigest()
    return h[:length]


def extract_month_from_filename(filename):
    """
    Пытается найти русское название месяца в имени файла.
    Возвращает (номер_месяца, нормализованное_имя) или (None, None).
    """
    basename = os.path.basename(filename)
    no_ext = os.path.splitext(basename)[0].lower()
    no_ext = re.sub(r'[^\w\s]', ' ', no_ext)
    no_ext = no_ext.replace('_', ' ')
    words = re.findall(r'\w+', no_ext)
    for w in words:
        parsed = morph.parse(w)[0]
        if parsed.normal_form in months:
            return months[parsed.normal_form], parsed.normal_form
    return None, None


def excel_column_to_index(column_letter):
    column_letter = column_letter.upper()
    index = 0
    for char in column_letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1


# -----------------------
# Логика "жёлтых" столбцов и объединённых ячеек
# -----------------------

def get_yellow_columns_pandas_style(sheet):
    """
    Ищет столбцы, где >=10 ячеек залиты цветом 'FFFFFF00' (ярко-жёлтый).
    Пропускает скрытые строки и слитые ячейки.
    Если найдены, также добавляет столбец перед первым жёлтым.
    Возвращает список 0-based индексов.
    """
    max_col = sheet.max_column
    max_row = sheet.max_row

    col_color_count = [0] * max_col

    for row_idx in range(1, max_row + 1):
        if sheet.row_dimensions[row_idx].hidden:
            continue
        row_cells = sheet[row_idx]
        for cell in row_cells:
            if isinstance(cell, MergedCell):
                continue
            if cell.fill and cell.fill.fgColor:
                fill_color = cell.fill.fgColor
                if fill_color.type == 'rgb' and fill_color.rgb == 'FFFFFF00':
                    col_i = cell.col_idx - 1
                    col_color_count[col_i] += 1

    yellow_columns = [i for i, count in enumerate(col_color_count) if count >= 10]

    if yellow_columns:
        first_yellow = yellow_columns[0]
        before_first = first_yellow - 1
        if before_first >= 0:
            yellow_columns.insert(0, before_first)

    return yellow_columns


def unmerge_horizontal_cells_in_memory(rows_data, merged_ranges, row_to_data_index, max_cols):
    """
    Если по горизонтали слито, и в верхней левой ячейке написано «того»,
    размазываем значение вправо.
    """
    for rng in merged_ranges:
        if rng.min_row == rng.max_row and rng.min_col < rng.max_col:
            sheet_row = rng.min_row
            if sheet_row not in row_to_data_index:
                continue
            row_idx = row_to_data_index[sheet_row]
            start_c = rng.min_col
            end_c = rng.max_col

            if start_c > max_cols:
                continue

            top_left_val = rows_data[row_idx][start_c - 1] if (start_c <= max_cols) else None

            if top_left_val and "того" in str(top_left_val).lower():
                for c in range(start_c, end_c + 1):
                    if c <= max_cols:
                        rows_data[row_idx][c - 1] = top_left_val


def fill_merged_cells_in_first_yellow_column_in_memory(rows_data, merged_ranges, first_yellow_idx,
                                                       row_to_data_index, max_cols):
    """
    Если вертикально слиты ячейки в первом жёлтом столбце, копируем значение
    верхней ячейки вниз по диапазону.
    """
    target_col = first_yellow_idx + 1  # openpyxl использует 1-based нумерацию
    for rng in merged_ranges:
        if rng.min_col == target_col and rng.max_col == target_col:
            top_left_row = rng.min_row
            top_left_val = None
            if top_left_row in row_to_data_index:
                row_idx = row_to_data_index[top_left_row]
                if target_col <= max_cols:
                    top_left_val = rows_data[row_idx][target_col - 1]
            for sheet_row in range(rng.min_row, rng.max_row + 1):
                if sheet_row in row_to_data_index:
                    row_idx = row_to_data_index[sheet_row]
                    if target_col <= max_cols:
                        rows_data[row_idx][target_col - 1] = top_left_val


def forward_fill_column_by_index_in_memory(rows_data, col_idx):
    """
    Простейший forward fill для указанного столбца в rows_data.
    """
    last_value = None
    for i in range(len(rows_data)):
        val = rows_data[i][col_idx]
        if val not in (None, ""):
            last_value = val
        else:
            if last_value is not None:
                rows_data[i][col_idx] = last_value


# -----------------------
# Очистка и фильтрация данных
# -----------------------

def fill_none_and_non_numeric_with_zero(df, columns_to_check=None):
    """
    Заполняет None и нечисловые значения нулями в указанных столбцах (по умолчанию Ruda, fRuda).
    """
    if columns_to_check is None:
        columns_to_check = ["Ruda", "fRuda"]

    filled_df = df.copy()
    filled_df[columns_to_check] = filled_df[columns_to_check].fillna(0)

    for col in columns_to_check:
        filled_df[col] = pd.to_numeric(filled_df[col], errors='coerce').fillna(0)

    return filled_df


def filter_columns_with_whole_number_sums(df, columns_to_check=None):
    """
    Удаляет строки, где сумма по Ruda, Cu, Ag, fRuda, fCu, fAg является целым числом и не превышает 180.
    """
    if columns_to_check is None:
        columns_to_check = ["Ruda", "Cu", "Ag", "fRuda", "fCu", "fAg"]

    filtered_df = df.copy()

    for index, row in filtered_df.iterrows():
        row_sum = 0
        has_non_numeric = False
        for col in columns_to_check:
            value = row[col]
            if pd.isna(value) or value == '' or value is None:
                value = 0
            try:
                row_sum += float(value)
            except (ValueError, TypeError):
                has_non_numeric = True
                break
        if (not has_non_numeric) and (row_sum == int(row_sum)) and (row_sum <= 180):
            filtered_df.drop(index, inplace=True)

    return filtered_df


# -----------------------
# Вычисления и агрегации
# -----------------------

def add_calculated_columns(df):
    cols = ['Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg']
    for c in cols:
        df[c] = df[c].replace([None, ''], np.nan)
        df[c] = df[c].astype(str).str.replace(',', '').str.strip()
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    df['diff-Ruda'] = abs(df['Ruda'] - df['fRuda'])
    df['diff-Cu'] = abs(df['Cu'] - df['fCu'])
    df['diff-Ag'] = abs(df['Ag'] - df['fAg'])

    df['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(df['Ruda'] != 0, 1 - df['diff-Ruda'] / df['Ruda'], 1.0),
        0, 100
    )
    df['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(df['Cu'] != 0, 1 - df['diff-Cu'] / df['Cu'], 1.0),
        0, 100
    )
    df['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(df['Ag'] != 0, 1 - df['diff-Ag'] / df['Ag'], 1.0),
        0, 100
    )

    df['%Cu'] = np.where(df['Ruda'] != 0, df['Cu'] / df['Ruda'] * 100, 0)
    df['%fCu'] = np.where(df['fRuda'] != 0, df['fCu'] / df['fRuda'] * 100, 0)

    df['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where(df['%Cu'] != 0, 1 - abs((df['%Cu'] - df['%fCu']) / df['%Cu']), 1.0),
        0, 100
    )

    df['p>0'] = np.where(df['Ruda'] > 0, 1, 0)
    df['f>0'] = np.where(df['fRuda'] > 0, 1, 0)
    return df


def group_by_block_and_month(df):
    sum_cols = ['Ruda', 'fRuda', 'Cu', 'fCu', 'Ag', 'fAg']
    grouped = df.groupby(['Panel', 'month'], as_index=False)[sum_cols].sum()

    grouped['Ruda-fRuda'] = grouped['Ruda'] - grouped['fRuda']
    grouped['Cu-fCu'] = grouped['Cu'] - grouped['fCu']
    grouped['Ag-fAg'] = grouped['Ag'] - grouped['fAg']

    grouped['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(grouped['Ruda'] != 0, 1 - abs(grouped['Ruda-fRuda'] / grouped['Ruda']), 1.0),
        0, 100
    )
    grouped['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(grouped['Cu'] != 0, 1 - abs(grouped['Cu-fCu'] / grouped['Cu']), 1.0),
        0, 100
    )
    grouped['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(grouped['Ag'] != 0, 1 - abs(grouped['Ag-fAg'] / grouped['Ag']), 1.0),
        0, 100
    )

    grouped['%Cu'] = np.where(grouped['Ruda'] != 0, grouped['Cu'] / grouped['Ruda'] * 100, 0)
    grouped['%fCu'] = np.where(grouped['fRuda'] != 0, grouped['fCu'] / grouped['fRuda'] * 100, 0)

    grouped['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where(grouped['%Cu'] != 0, 1 - abs((grouped['%Cu'] - grouped['%fCu']) / grouped['%Cu']), 1.0),
        0, 100
    )

    grouped['p>0'] = np.where(grouped['Ruda'] > 0, 1, 0)
    grouped['f>0'] = np.where(grouped['fRuda'] > 0, 1, 0)
    return grouped


def group_by_block(df):
    sum_cols = ['Ruda', 'fRuda', 'Cu', 'fCu', 'Ag', 'fAg']
    grouped = df.groupby(['Panel'], as_index=False)[sum_cols].sum()

    grouped['Ruda-fRuda'] = grouped['Ruda'] - grouped['fRuda']
    grouped['Cu-fCu'] = grouped['Cu'] - grouped['fCu']
    grouped['Ag-fAg'] = grouped['Ag'] - grouped['fAg']

    grouped['%Cu'] = np.where(grouped['Ruda'] != 0, grouped['Cu'] / grouped['Ruda'] * 100, 0)
    grouped['%fCu'] = np.where(grouped['fRuda'] != 0, grouped['fCu'] / grouped['fRuda'] * 100, 0)

    grouped['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(grouped['Ruda'] != 0, 1 - abs(grouped['Ruda-fRuda'] / grouped['Ruda']), 1.0),
        0, 100
    )
    grouped['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(grouped['Cu'] != 0, 1 - abs(grouped['Cu-fCu'] / grouped['Cu']), 1.0),
        0, 100
    )
    grouped['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(grouped['Ag'] != 0, 1 - abs(grouped['Ag-fAg'] / grouped['Ag']), 1.0),
        0, 100
    )
    grouped['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where(grouped['%Cu'] != 0, 1 - abs((grouped['%Cu'] - grouped['%fCu']) / grouped['%Cu']), 1.0),
        0, 100
    )

    grouped['p>0'] = np.where(grouped['Ruda'] > 0, 1, 0)
    grouped['f>0'] = np.where(grouped['fRuda'] > 0, 1, 0)
    return grouped


def mean_values_groupped_by_month(df):
    needed = ['1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']
    grouped = df.groupby('month', as_index=False)[needed].mean()
    return grouped


# -----------------------
# Окончательная отчётность
# -----------------------

def generate_report_with_charts(folder_path, full_df):
    """
    Объединяет результаты, добавляет вычисляемые столбцы,
    группирует данные и записывает итоговый отчёт в Excel.
    """
    full_df = add_calculated_columns(full_df)

    block_and_month_aggregated_df = group_by_block_and_month(full_df)
    block_aggregated_df = group_by_block(full_df)
    monthly_avg_df = mean_values_groupped_by_month(full_df)

    t = datetime.now().microsecond
    output_file_name = f"_report_{os.path.basename(folder_path.strip('/\\'))}{t}.xlsx"
    output_file = os.path.join(folder_path, output_file_name)

    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        full_df.to_excel(writer, sheet_name="По штрекам", index=False)
        block_and_month_aggregated_df.to_excel(writer, sheet_name="Сумм. по панелям", index=False)
        block_aggregated_df.to_excel(writer, sheet_name="Сумм. по блокам", index=False)
        monthly_avg_df.to_excel(writer, sheet_name="Средн. по месяцам", index=False)

        workbook = writer.book
        worksheet = workbook.add_worksheet("Some Charts")
        # Здесь можно добавить код для построения графиков, если потребуется

    print(f"\n==> Итоговый отчёт сохранён: {output_file}")


# -----------------------
# Функция назначения блоков (как в оригинальном алгоритме)
# -----------------------

def assign_block_numbers(data, col_idx):
    """
    Для DataFrame data с уже выбранными "жёлтыми" столбцами.
    col_idx – индекс столбца (обычно столбца непосредственно перед первым жёлтым),
    по которому определяется наличие слова "того" (делимитер).

    Алгоритм:
      - Если строка содержит "того", в соответствующем элементе списка block_names записывается None,
        а текущая панель сбрасывается.
      - Иначе, если текущая панель ещё не установлена, берётся значение из второй колонки данной строки.
      - В конце создаётся колонка Block, а строки с None отбрасываются.
    """
    is_delimiter = data.iloc[:, col_idx].astype(str).str.lower().str.contains('того', na=False)
    panel_names = []
    current_panel = None

    for idx, is_delim in enumerate(is_delimiter):
        if is_delim:
            panel_names.append(None)
            current_panel = None
        else:
            if current_panel is None:
                current_panel = data.iloc[idx, 1]
            panel_names.append(current_panel)

    data = data.copy()
    data['Block'] = panel_names
    data = data.loc[data['Block'].notna()].copy()
    return data


def parse_string(value):
    """
    Преобразует строку: сохраняет первый символ, остальные приводит к нижнему регистру
    и добавляет пробел между цифрой и последующей буквой.
    """
    if isinstance(value, int):
        return value
    if not isinstance(value, str):
        value = str(value)
    value = value.strip()
    if not value:
        return value
    formatted_value = value[0] + value[1:].lower() if len(value) > 1 else value
    formatted_value = re.sub(r'(?<=\d)(?=[A-Za-zА-Яа-яЁё])', ' ', formatted_value)
    return formatted_value


# -----------------------
# Основная функция: объединённая загрузка
# -----------------------

def load_excel_data_with_flex(folder_path, tip=1, max_rows=500):
    """
    1) Converts .xls/.xlsb files to .xlsx if needed.
    2) For each .xlsx file:
         - Opens the workbook using openpyxl (taking hidden rows, merged cells,
           yellow columns, etc. into account).
         - Builds a 2D list (rows_data) of visible rows with corrections for merged cells.
         - Converts the data into a DataFrame and selects only the "yellow" columns.
         - If the sheet name contains "ОГР":
              • Adjusts column names as follows:
                – If the number of columns equals 9, it is assumed that the "Horizont" column exists.
                – If the number of columns equals 8, an empty "Horizont" column is inserted on the left.
              • Then calls assign_block_numbers() to generate the Block column.
              • Finally, appends the string " ОГР" to every Panel value.
         - Otherwise:
              • Renames the columns according to the standard (9 columns) and sets Block equal to Panel.
         - Additional filtering and processing are then applied.
    3) All data are combined into a single DataFrame, which is then aggregated and used to generate a report in Excel.
    """
    from glob import glob
    import os
    import openpyxl
    import pandas as pd
    # Assuming the helper functions (convert_xls_to_xlsx_with_formatting, convert_xlsb_to_xlsx_with_formatting,
    # get_yellow_columns_pandas_style, unmerge_horizontal_cells_in_memory, fill_merged_cells_in_first_yellow_column_in_memory,
    # fill_none_and_non_numeric_with_zero, filter_columns_with_whole_number_sums, assign_block_numbers)
    # and extract_month_from_filename are defined elsewhere in the script.

    # 1) Conversion: Convert .xls and .xlsb files to .xlsx if needed.
    xls_files = glob(os.path.join(folder_path, "*.xls"))
    for xls in xls_files:
        out_xlsx = os.path.splitext(xls)[0] + ".xlsx"
        if not os.path.exists(out_xlsx):
            convert_xls_to_xlsx_with_formatting(xls, out_xlsx)
    xlsb_files = glob(os.path.join(folder_path, "*.xlsb"))
    for xlsb in xlsb_files:
        out_xlsx = os.path.splitext(xlsb)[0] + ".xlsx"
        if not os.path.exists(out_xlsx):
            convert_xlsb_to_xlsx_with_formatting(xlsb, out_xlsx)

    # 2) Reading .xlsx files.
    xlsx_files = glob(os.path.join(folder_path, "*.xlsx"))
    final_columns = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'Ag',
                     'fRuda', 'fCu', 'fAg', 'Uchastok', 'month', 'Block']
    final_df = pd.DataFrame(columns=final_columns)

    for file_path in xlsx_files:
        file_name = os.path.basename(file_path)
        month_num, month_name = extract_month_from_filename(file_name)
        if not month_num:
            print(f"Нет месяца в названии файла - пропускаю: {file_name}")
            continue

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        except Exception as e:
            print(f"Не удалось открыть {file_path}: {e}")
            continue

        visible_sheets = [sh for sh in wb.sheetnames if wb[sh].sheet_state == "visible"]
        if len(visible_sheets) > 1:
            visible_sheets = visible_sheets[1:]  # as in the original code

        for sheet_name in visible_sheets:
            ws = wb[sheet_name]
            if ws.max_row == 0 or ws.max_column == 0:
                print(f"Лист '{sheet_name}' пуст в файле {file_name}")
                continue

            print(f"\nОбработка {file_name} => лист '{sheet_name}' => месяц: {month_name}")

            max_row_to_process = min(ws.max_row, max_rows)
            merged_ranges = list(ws.merged_cells.ranges)
            yellow_cols = get_yellow_columns_pandas_style(ws)
            if not yellow_cols:
                print(f"Нет 'жёлтых' столбцов в листе '{sheet_name}' => пропускаю.")
                continue
            first_yellow_idx = yellow_cols[0]

            # Build rows_data, skipping hidden rows.
            visible_rows = []
            rows_data = []
            for row_num in range(1, max_row_to_process + 1):
                if ws.row_dimensions[row_num].hidden:
                    continue
                row_cells = ws[row_num]
                row_vals = []
                for cell in row_cells:
                    if isinstance(cell, openpyxl.cell.MergedCell):
                        row_vals.append(None)
                    else:
                        row_vals.append(cell.value)
                rows_data.append(row_vals)
                visible_rows.append(row_num)

            if not rows_data:
                print(f"Нет данных после пропуска скрытых строк: {sheet_name}")
                continue

            row_to_data_index = {r: i for i, r in enumerate(visible_rows)}
            max_col_count = max(len(r) for r in rows_data)

            unmerge_horizontal_cells_in_memory(rows_data, merged_ranges, row_to_data_index, max_col_count)
            fill_merged_cells_in_first_yellow_column_in_memory(rows_data, merged_ranges, first_yellow_idx,
                                                               row_to_data_index, max_col_count)
            # Optionally: forward_fill_column_by_index_in_memory(rows_data, first_yellow_idx)

            header_row = [str(x) if x else "" for x in rows_data[0]]
            data_body = rows_data[1:]
            if not data_body:
                print(f"Пустое тело данных (лист {sheet_name}). Пропускаю.")
                continue

            data = pd.DataFrame(data_body, columns=header_row)
            data = data.dropna(how='all')

            max_index = data.shape[1] - 1
            valid_yellows = [c for c in yellow_cols if 0 <= c <= max_index]
            if not valid_yellows:
                print(f"Жёлтые столбцы вне диапазона в листе {sheet_name}. Пропускаю.")
                continue
            data = data.iloc[:, valid_yellows]

            # Rename columns based on sheet name.
            if "ОГР" in sheet_name.upper():
                # For ОГР sheets:
                # If there are 9 columns, assume 'Horizont' exists.
                # If there are 8, insert an empty 'Horizont' column on the left.
                if len(data.columns) == 9:
                    std_cols = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'fRuda', 'fCu', 'Ag', 'fAg']
                    data.columns = std_cols
                elif len(data.columns) == 8:
                    std_cols = ['Shtrek', 'Panel', 'Ruda', 'Cu', 'fRuda', 'fCu', 'Ag', 'fAg']
                    data.columns = std_cols
                    data.insert(0, 'Horizont', '')
                else:
                    print(f"Несоответствие числа столбцов в листе '{sheet_name}' (ОГР). Пропускаю.")
                    continue
            else:
                if tip == 1:
                    std_cols = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg']
                else:
                    std_cols = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'fRuda', 'fCu', 'Ag', 'fAg']
                if len(data.columns) == len(std_cols):
                    data.columns = std_cols
                else:
                    print(f"Несоответствие числа столбцов в листе '{sheet_name}'. Пропускаю.")
                    continue

            # Remove header duplicate row if necessary.
            if len(data) > 1:
                data = data.iloc[1:]
            else:
                data = data.iloc[0:0]

            # Filter out rows with empty Panel.
            data = data[data['Panel'].notnull() & (data['Panel'] != '')]

            # Remove rows where Panel, Shtrek, or Horizont contain "того" or "итог".
            for col_name in ['Panel', 'Shtrek', 'Horizont']:
                if col_name in data.columns:
                    data = data[~data[col_name].astype(str).str.lower().str.contains('того', na=False)]
                    data = data[~data[col_name].astype(str).str.lower().str.contains('итог', na=False)]

            data = fill_none_and_non_numeric_with_zero(data)
            data = filter_columns_with_whole_number_sums(data)

            if 'Panel' in data.columns:
                data = data[data['Panel'].notnull() & (data['Panel'] != '')]

            # Add additional columns.
            data['Uchastok'] = sheet_name
            data['month'] = month_name

            # Ensure all final columns exist.
            for col in final_columns:
                if col not in data.columns:
                    data[col] = ''

            data = data[~((data['Ruda'] == 0) & (data['fRuda'] == 0))]
            data = data[~(data['fAg'] == 'кг')]

            # Form the Block column.
            if "ОГР" in sheet_name.upper():
                # For ОГР sheets, use the original algorithm.
                col_before_yellow_idx = valid_yellows[0] - 1
                if col_before_yellow_idx < 0:
                    col_before_yellow_idx = 0
                data = assign_block_numbers(data, col_before_yellow_idx)
                # Append " ОГР" to every Panel name.
                data["Panel"] = "ОГР " + data["Panel"].astype(str)
            else:
                # For non-ОГР sheets, simply copy Panel to Block.
                data['Block'] = data['Panel'].astype(str)

            data = data[final_columns]
            final_df = pd.concat([final_df, data], ignore_index=True)

    final_df = final_df[~((final_df['Ruda'] == 0) & (final_df['fRuda'] == 0))]
    final_df['month_N'] = final_df['month'].map(lambda m: months.get(m.lower(), 0))
    final_df = final_df.sort_values(by=['month_N', 'Block'])
    final_df.drop(columns=['month_N'], inplace=True)

    generate_report_with_charts(folder_path, final_df)
    return final_df

# -----------------------
# Пример использования
# -----------------------

if __name__ == "__main__":
    tip = 1
    # Укажите директорию с файлами Excel
    directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Жиланды ИПГ 2024'
    final_data = load_excel_data_with_flex(directory, tip, max_rows=200)
    print(final_data.head())
