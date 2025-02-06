import os
from datetime import datetime
import pandas as pd
from glob import glob
import re
import hashlib
import warnings
import numpy as np

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell

warnings.filterwarnings('ignore', category=FutureWarning)

try:
    import win32com.client as win32
except ImportError:
    win32 = None  # на non-Windows или без установленного Excel

from pymorphy2 import MorphAnalyzer

# ----------------------------
# Словарь для русских месяцев
# ----------------------------
morph = MorphAnalyzer()
months = {
    'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5, 'июнь': 6,
    'июль': 7, 'август': 8, 'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12
}

COL_LIMIT = 94

###############################################################################
#                         Conversion Functions
###############################################################################

def convert_xls_to_xlsx_with_formatting(xls_file_path, xlsx_file_path):
    """
    Конвертирует .xls в .xlsx, сохраняя форматирование (требует Windows + Excel).
    """
    if win32 is None:
        print("win32com недоступен; не могу конвертировать .xls.")
        return
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(xls_file_path)
        workbook.SaveAs(xlsx_file_path, FileFormat=51)  # 51 => .xlsx
        workbook.Close()
    except Exception as e:
        print(f"Ошибка при конвертации {xls_file_path}: {e}")
    finally:
        excel.Quit()


def convert_xlsb_to_xlsx_with_formatting(xlsb_file_path, xlsx_file_path):
    """
    Конвертирует .xlsb в .xlsx, сохраняя форматирование (требует Windows + Excel).
    """
    if win32 is None:
        print("win32com недоступен; не могу конвертировать .xlsb.")
        return
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(xlsb_file_path)
        workbook.SaveAs(xlsx_file_path, FileFormat=51)  # 51 => .xlsx
        workbook.Close()
    except Exception as e:
        print(f"Ошибка при конвертации {xlsb_file_path}: {e}")
    finally:
        excel.Quit()


###############################################################################
#                         Support / Helper Functions
###############################################################################

def generate_hash(s, length=2):
    if not isinstance(s, str):
        s = str(s)
    h = hashlib.sha256(s.encode()).hexdigest()
    return h[:length]


def extract_month_from_filename(filename):
    """
    Пытается найти русское название месяца в имени файла.
    Возвращает (номер_месяца, нормализованное_имя) или (None, None).
    """
    basename = os.path.basename(filename)
    no_ext = os.path.splitext(basename)[0].lower()
    no_ext = re.sub(r'[^\w\s]', ' ', no_ext)
    no_ext = no_ext.replace('_', ' ')
    words = re.findall(r'\w+', no_ext)
    for w in words:
        parsed = morph.parse(w)[0]
        if parsed.normal_form in months:
            return months[parsed.normal_form], parsed.normal_form
    return None, None


def excel_column_to_index(column_letter):
    column_letter = column_letter.upper()
    index = 0
    for char in column_letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1


###############################################################################
#             Поиск «жёлтых» столбцов (и колонки до первого жёлтого)
###############################################################################

def get_yellow_columns_pandas_style(sheet):
    """
    Ищем столбцы, где >=10 ячеек залиты цветом "FFFFFF00" (ярко-жёлтый).
    Пропускаем скрытые строки и слитые ячейки.
    Возвращает список 0-based индексов.
    Если найдены, добавляем «столбец перед первым жёлтым» в начало списка.
    """
    max_col = sheet.max_column
    max_row = sheet.max_row

    col_color_count = [0] * max_col

    for row_idx in range(1, max_row + 1):
        # пропускаем скрытые строки
        if sheet.row_dimensions[row_idx].hidden:
            continue
        row_cells = sheet[row_idx]
        for cell in row_cells:
            if isinstance(cell, MergedCell):
                continue
            if cell.fill and cell.fill.fgColor:
                fill_color = cell.fill.fgColor
                # ищем именно 'FFFFFF00'
                if fill_color.type == 'rgb' and fill_color.rgb == 'FFFFFF00':
                    col_i = cell.col_idx - 1  # 0-based
                    col_color_count[col_i] += 1

    # столбцы, где >=10 жёлтых ячеек
    yellow_columns = [i for i, count in enumerate(col_color_count) if count >= 10]

    if yellow_columns:
        first_yellow = yellow_columns[0]
        before_first = first_yellow - 1
        if before_first >= 0:
            yellow_columns.insert(0, before_first)

    return yellow_columns


###############################################################################
#            Unmerge & Fill Logic, если в «первом жёлтом» merge и т.д.
###############################################################################

def unmerge_horizontal_cells_in_memory(rows_data, merged_ranges, row_to_data_index, max_cols):
    """
    Исходная логика: если по горизонтали слито,
    и в верхней левой ячейке написано «того», размазываем её значение вправо.
    """
    for rng in merged_ranges:
        # если merge в одну строку (min_row == max_row) и несколько столбцов
        if rng.min_row == rng.max_row and rng.min_col < rng.max_col:
            sheet_row = rng.min_row
            if sheet_row not in row_to_data_index:
                continue
            row_idx = row_to_data_index[sheet_row]
            start_c = rng.min_col
            end_c   = rng.max_col

            if start_c > max_cols:
                continue

            top_left_val = rows_data[row_idx][start_c - 1] if (start_c <= max_cols) else None

            if top_left_val and "того" in str(top_left_val).lower():
                for c in range(start_c, end_c + 1):
                    if c <= max_cols:
                        rows_data[row_idx][c - 1] = top_left_val


def fill_merged_cells_in_first_yellow_column_in_memory(rows_data, merged_ranges, first_yellow_idx,
                                                       row_to_data_index, max_cols):
    """
    Если merge вертикальный по «первому жёлтому» столбцу, копируем значение верхней ячейки вниз.
    """
    target_col = first_yellow_idx + 1  # для openpyxl это 1-based
    for rng in merged_ranges:
        # проверяем что слито ровно в 1 столбец
        if rng.min_col == target_col and rng.max_col == target_col:
            top_left_row = rng.min_row
            top_left_val = None
            if top_left_row in row_to_data_index:
                row_idx = row_to_data_index[top_left_row]
                if target_col <= max_cols:
                    top_left_val = rows_data[row_idx][target_col - 1]

            for sheet_row in range(rng.min_row, rng.max_row + 1):
                if sheet_row in row_to_data_index:
                    row_idx = row_to_data_index[sheet_row]
                    if target_col <= max_cols:
                        rows_data[row_idx][target_col - 1] = top_left_val


def forward_fill_column_by_index_in_memory(rows_data, col_idx):
    """
    Простейший ffill в «сыром» массиве rows_data по определённому столбцу.
    """
    last_value = None
    for i in range(len(rows_data)):
        val = rows_data[i][col_idx]
        if val not in (None, ""):
            last_value = val
        else:
            if last_value is not None:
                rows_data[i][col_idx] = last_value


###############################################################################
#                             Filtering Logic
###############################################################################

def fill_none_and_non_numeric_with_zero(df, columns_to_check=None):
    """
    Заполняет None и нечисловые значения нулями в указанных столбцах (по умолчанию Ruda,fRuda).
    """
    if columns_to_check is None:
        columns_to_check = ["Ruda", "fRuda"]

    filled_df = df.copy()
    filled_df[columns_to_check] = filled_df[columns_to_check].fillna(0)

    for col in columns_to_check:
        filled_df[col] = pd.to_numeric(filled_df[col], errors='coerce').fillna(0)

    return filled_df


def filter_columns_with_whole_number_sums(df, columns_to_check=None):
    """
    Старое правило: если сумма по Ruda,Cu,Ag,fRuda,fCu,fAg является целым числом <=180, удаляем строку.
    """
    if columns_to_check is None:
        columns_to_check = ["Ruda", "Cu", "Ag", "fRuda", "fCu", "fAg"]

    filtered_df = df.copy()
    for index, row in filtered_df.iterrows():
        row_sum = 0
        has_non_numeric = False
        for col in columns_to_check:
            value = row[col]
            if pd.isna(value) or value == '' or value is None:
                value = 0
            try:
                row_sum += float(value)
            except (ValueError, TypeError):
                has_non_numeric = True
                break
        if (not has_non_numeric) and (row_sum == int(row_sum)) and (row_sum <= 180):
            filtered_df.drop(index, inplace=True)

    return filtered_df


###############################################################################
#               Calculations & Aggregations (unchanged logic)
###############################################################################

def add_calculated_columns(df):
    cols = ['Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg']
    for c in cols:
        df[c] = df[c].replace([None, ''], np.nan)
        df[c] = df[c].astype(str).str.replace(',', '').str.strip()
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    df['diff-Ruda'] = abs(df['Ruda'] - df['fRuda'])
    df['diff-Cu']   = abs(df['Cu']   - df['fCu'])
    df['diff-Ag']   = abs(df['Ag']   - df['fAg'])

    df['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(df['Ruda'] != 0, 1 - df['diff-Ruda'] / df['Ruda'], 1.0),
        0, 100
    )
    df['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(df['Cu'] != 0, 1 - df['diff-Cu'] / df['Cu'], 1.0),
        0, 100
    )
    df['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(df['Ag'] != 0, 1 - df['diff-Ag'] / df['Ag'], 1.0),
        0, 100
    )

    df['%Cu']  = np.where(df['Ruda']  != 0, df['Cu']  / df['Ruda']  * 100, 0)
    df['%fCu'] = np.where(df['fRuda'] != 0, df['fCu'] / df['fRuda'] * 100, 0)

    df['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where(df['%Cu'] != 0, 1 - abs((df['%Cu'] - df['%fCu']) / df['%Cu']), 1.0),
        0, 100
    )

    df['p>0'] = np.where(df['Ruda']  > 0, 1, 0)
    df['f>0'] = np.where(df['fRuda'] > 0, 1, 0)
    return df


def group_by_block_and_month(df):
    sum_cols = ['Ruda', 'fRuda', 'Cu', 'fCu', 'Ag', 'fAg']
    grouped = df.groupby(['Panel', 'month'], as_index=False)[sum_cols].sum()

    grouped['Ruda-fRuda'] = grouped['Ruda'] - grouped['fRuda']
    grouped['Cu-fCu']     = grouped['Cu']   - grouped['fCu']
    grouped['Ag-fAg']     = grouped['Ag']   - grouped['fAg']

    grouped['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(grouped['Ruda'] != 0, 1 - abs(grouped['Ruda-fRuda'] / grouped['Ruda']), 1.0),
        0, 100
    )
    grouped['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(grouped['Cu'] != 0, 1 - abs(grouped['Cu-fCu'] / grouped['Cu']), 1.0),
        0, 100
    )
    grouped['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(grouped['Ag'] != 0, 1 - abs(grouped['Ag-fAg'] / grouped['Ag']), 1.0),
        0, 100
    )

    grouped['%Cu']  = np.where(grouped['Ruda']  != 0, grouped['Cu'] / grouped['Ruda']  * 100, 0)
    grouped['%fCu'] = np.where(grouped['fRuda'] != 0, grouped['fCu']/grouped['fRuda'] * 100, 0)

    grouped['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where(grouped['%Cu'] != 0, 1 - abs((grouped['%Cu'] - grouped['%fCu']) / grouped['%Cu']), 1.0),
        0, 100
    )

    grouped['p>0'] = np.where(grouped['Ruda']  > 0, 1, 0)
    grouped['f>0'] = np.where(grouped['fRuda'] > 0, 1, 0)
    return grouped


def group_by_block(df):
    sum_cols = ['Ruda', 'fRuda', 'Cu', 'fCu', 'Ag', 'fAg']
    grouped = df.groupby(['Panel'], as_index=False)[sum_cols].sum()

    grouped['Ruda-fRuda'] = grouped['Ruda'] - grouped['fRuda']
    grouped['Cu-fCu']     = grouped['Cu']   - grouped['fCu']
    grouped['Ag-fAg']     = grouped['Ag']   - grouped['fAg']

    grouped['%Cu']  = np.where(grouped['Ruda']  != 0, grouped['Cu']  / grouped['Ruda']  * 100, 0)
    grouped['%fCu'] = np.where(grouped['fRuda'] != 0, grouped['fCu'] / grouped['fRuda'] * 100, 0)

    grouped['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(grouped['Ruda'] != 0, 1 - abs(grouped['Ruda-fRuda'] / grouped['Ruda']), 1.0),
        0, 100
    )
    grouped['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(grouped['Cu'] != 0, 1 - abs(grouped['Cu-fCu'] / grouped['Cu']), 1.0),
        0, 100
    )
    grouped['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(grouped['Ag'] != 0, 1 - abs(grouped['Ag-fAg'] / grouped['Ag']), 1.0),
        0, 100
    )
    grouped['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where(grouped['%Cu'] != 0, 1 - abs((grouped['%Cu'] - grouped['%fCu']) / grouped['%Cu']), 1.0),
        0, 100
    )

    grouped['p>0'] = np.where(grouped['Ruda']  > 0, 1, 0)
    grouped['f>0'] = np.where(grouped['fRuda'] > 0, 1, 0)
    return grouped


def mean_values_groupped_by_month(df):
    needed = ['1-(Ruda-fRuda)/Ruda','1-(Cu-fCu)/Cu','1-(Ag-fAg)/Ag','1-(%Cu-%fCu)/%Cu']
    grouped = df.groupby('month', as_index=False)[needed].mean()
    return grouped


###############################################################################
#       Final Reporting
###############################################################################

def generate_report_with_charts(folder_path, full_df):
    """
    Объединяем результаты, пишем в Excel.
    """
    full_df = add_calculated_columns(full_df)

    block_and_month_aggregated_df = group_by_block_and_month(full_df)
    block_aggregated_df           = group_by_block(full_df)
    monthly_avg_df                = mean_values_groupped_by_month(full_df)

    t = datetime.now().microsecond
    output_file_name = f"_report_{os.path.basename(folder_path.strip('/\\'))}{t}.xlsx"
    output_file = os.path.join(folder_path, output_file_name)

    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        full_df.to_excel(writer, sheet_name="По штрекам", index=False)
        block_and_month_aggregated_df.to_excel(writer, sheet_name="Сумм. по панелям", index=False)
        block_aggregated_df.to_excel(writer, sheet_name="Сумм. по блокам", index=False)
        monthly_avg_df.to_excel(writer, sheet_name="Средн. по месяцам", index=False)

        workbook  = writer.book
        worksheet = workbook.add_worksheet("Some Charts")
        # по необходимости формируем графики xlsxwriter...

    print(f"\n==> Итоговый отчёт сохранён: {output_file}")


###############################################################################
#       Главная функция load_excel_data_with_flex
###############################################################################

def load_excel_data_with_flex(folder_path, tip=1, max_rows=500):
    """
    1) Конвертируем .xls / .xlsb -> .xlsx при необходимости.
    2) Читаем каждую .xlsx-книгу:
       - Пропускаем файлы, где не нашли месяц
       - Пропускаем скрытые листы, берём со 2-го листа если их много (как старый код)
       - Считываем скрытые строки (openpyxl), unmerge горизонталь, и т.д.
       - Находим «жёлтые» столбцы и добавляем "столбец перед первым"
       - Вырезаем из DataFrame только эти «жёлтые» столбцы: `data = data.iloc[:, valid_yellows]`
       - Переименовываем столбцы в зависимости от (ОГР или tip=1/2)
       - Удаляем первые строки, строки "того", и т.д.
       - **Block = Panel** (вместо assign_block_numbers)
       - Фильтруем и складываем в общий final_df
    3) Генерируем общий отчёт.
    """
    # 1) конвертация (как раньше)
    xls_files = glob(os.path.join(folder_path, "*.xls"))
    for xls in xls_files:
        out_xlsx = os.path.splitext(xls)[0] + ".xlsx"
        if not os.path.exists(out_xlsx):
            convert_xls_to_xlsx_with_formatting(xls, out_xlsx)

    xlsb_files = glob(os.path.join(folder_path, "*.xlsb"))
    for xlsb in xlsb_files:
        out_xlsx = os.path.splitext(xlsb)[0] + ".xlsx"
        if not os.path.exists(out_xlsx):
            convert_xlsb_to_xlsx_with_formatting(xlsb, out_xlsx)

    # 2) чтение
    xlsx_files = glob(os.path.join(folder_path, "*.xlsx"))
    final_columns = ['Horizont','Panel','Shtrek','Ruda','Cu','Ag','fRuda','fCu','fAg','Uchastok','month','Block']
    final_df = pd.DataFrame(columns=final_columns)

    for file_path in xlsx_files:
        file_name = os.path.basename(file_path)
        month_num, month_name = extract_month_from_filename(file_name)
        if not month_num:
            print(f"-> -> ->  нет месяца в названии файла - пропускаю: {file_name}")
            continue

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        except:
            print(f"Не удалось открыть {file_path}")
            continue

        visible_sheets = [sh for sh in wb.sheetnames if wb[sh].sheet_state == "visible"]
        if len(visible_sheets) > 1:
            visible_sheets = visible_sheets[1:]  # как в старом коде

        for sheet_name in visible_sheets:
            ws = wb[sheet_name]
            if ws.max_row == 0 or ws.max_column == 0:
                print(f"Лист '{sheet_name}' пуст в файле {file_name}")
                continue

            print(f"\n>>>>>>>>>>>>>>>> {file_name} => лист '{sheet_name}' => месяц: {month_name}")

            max_row_to_process = min(ws.max_row, max_rows)

            merged_ranges = list(ws.merged_cells.ranges)
            yellow_cols   = get_yellow_columns_pandas_style(ws)
            if not yellow_cols:
                print(f"Нет 'жёлтых' столбцов в листе '{sheet_name}' => пропускаю.")
                continue
            first_yellow_idx = yellow_cols[0]

            # Собираем raw data, пропуская скрытые строки
            visible_rows = []
            rows_data = []
            for row_num in range(1, max_row_to_process+1):
                if ws.row_dimensions[row_num].hidden:
                    continue
                row_cells = ws[row_num]
                row_vals = []
                for cell in row_cells:
                    if isinstance(cell, MergedCell):
                        row_vals.append(None)
                    else:
                        row_vals.append(cell.value)
                rows_data.append(row_vals)
                visible_rows.append(row_num)

            if not rows_data:
                print(f"После пропуска скрытых строк данных нет: {sheet_name}. Пропускаю.")
                continue

            row_to_data_index = { r: i for i, r in enumerate(visible_rows) }
            max_col_count = max(len(r) for r in rows_data)

            # unmerge по горизонтали если "того"
            unmerge_horizontal_cells_in_memory(rows_data, merged_ranges, row_to_data_index, max_col_count)
            # fill merged cells в первом жёлтом столбце
            fill_merged_cells_in_first_yellow_column_in_memory(
                rows_data, merged_ranges, first_yellow_idx,
                row_to_data_index, max_col_count
            )
            # forward fill при желании
            # forward_fill_column_by_index_in_memory(rows_data, first_yellow_idx)

            # Превращаем в DataFrame
            header_row = [str(x) if x else "" for x in rows_data[0]]
            data_body  = rows_data[1:]
            if not data_body:
                print(f"Пустое тело данных (sheet {sheet_name}). Пропускаю.")
                continue

            data = pd.DataFrame(data_body, columns=header_row)
            data = data.dropna(how='all')

            # ВАЖНО: берём только жёлтые столбцы
            max_index = data.shape[1] - 1
            valid_yellows = [c for c in yellow_cols if 0 <= c <= max_index]
            if not valid_yellows:
                print(f"Столбцы-жёлтые вне диапазона? {sheet_name} Пропуск.")
                continue
            data = data.iloc[:, valid_yellows]

            # forward fill первых 3 столбцов (часто нужно, как в старом коде)
            # if data.shape[1] > 0:
            #     data.iloc[:,0] = data.iloc[:,0].ffill()
            # if data.shape[1] > 1:
            #     data.iloc[:,1] = data.iloc[:,1].ffill()
            # if data.shape[1] > 2:
            #     data.iloc[:,2] = data.iloc[:,2].ffill()

            # Логика переименования столбцов (как в старом коде)
            if 'ОГР' in sheet_name.upper():
                # Ожидаем 8 столбцов
                ogr_cols = ['Shtrek','Panel','Ruda','Cu','fRuda','fCu','Ag','fAg']
                if len(data.columns) == len(ogr_cols):
                    data.columns = ogr_cols
                    # добавляем пустой Horizont
                    data['Horizont'] = ''
                else:
                    print(f"Warning: Лист «{sheet_name}» => несоответствие числа столбцов (ОГР).")
                    continue
            else:
                # tip=1 => ожидаем 10 столбцов
                # tip=2 => ожидаем 9 столбцов
                if tip == 1:
                    std_cols = ['Horizont','Panel','Shtrek','Ruda','Cu','Ag','fRuda','fCu','fAg']
                else:
                    std_cols = ['Horizont','Panel','Shtrek','Ruda','Cu','fRuda','fCu','Ag','fAg']

                if len(data.columns) == len(std_cols):
                    data.columns = std_cols
                else:
                    print(f"Warning: Лист «{sheet_name}» => несоответствие числа столбцов.")
                    continue


            # # old code: col_before_yellow = yellow_cols[0] - 1 => used to parse 'ИТОГО', etc.
            # # but we have that logic in unmerge. We'll do the same:
            # col_before_yellow_idx = valid_yellows[0] - 1
            # if col_before_yellow_idx < 0:
            #     col_before_yellow_idx = 0
            #
            # is_togo = data.iloc[:, col_before_yellow_idx].astype(str).str.lower().str.contains('того', na=False)
            # # Remove the rows that contain 'того'
            # data = data[~is_togo]

            data = data[data['Panel'].notnull() & (data['Panel'] != '')]

            # # Удаляем первую строку (если надо)
            # if len(data) > 1:
            #     data = data.iloc[1:]
            # else:
            #     data = data.iloc[0:0]

            # Вместо старой логики assign_block_numbers(...) => Block = Panel
            if 'Block' in data.columns and 'Panel' in data.columns:
                data['Block'] = data['Panel'].astype(str)

            # Убираем строки, где в Panel/Shtrek/Horizont написано "того"
            for col_name in ['Panel','Shtrek','Horizont']:
                if col_name in data.columns:
                    data = data[~data[col_name].astype(str).str.lower().str.contains('того', na=False)]
                    data = data[~data[col_name].astype(str).str.lower().str.contains('итог', na=False)]


            # Заполняем нулями Ruda/fRuda
            data = fill_none_and_non_numeric_with_zero(data)

            # Фильтрация целочисленных сумм
            data = filter_columns_with_whole_number_sums(data)

            # Пропуск, если Shtrek или Panel пустые (как в старом коде)
            # if 'Shtrek' in data.columns:
            #     data = data[data['Shtrek'].notnull() & (data['Shtrek'] != '')]
            if 'Panel' in data.columns:
                data = data[data['Panel'].notnull() & (data['Panel'] != '')]

            # Добавляем Uchastok, month
            data['Uchastok'] = sheet_name
            data['month']    = month_name

            # Гарантируем, что все нужные столбцы есть
            for col in final_columns:
                if col not in data.columns:
                    data[col] = ''

            # Удаляем строки, где Ruda=0 и fRuda=0
            data = data[~((data['Ruda'] == 0) & (data['fRuda'] == 0))]

            # Удаляем fAg == 'кг'
            data = data[~(data['fAg'] == 'кг')]

            # Переупорядочиваем столбцы
            data = data[final_columns]

            final_df = pd.concat([final_df, data], ignore_index=True)

    # Финальная фильтрация
    final_df = final_df[~((final_df['Ruda'] == 0) & (final_df['fRuda'] == 0))]

    # Сортируем по месяцу + Block
    final_df['month_N'] = final_df['month'].map(months)
    final_df = final_df.sort_values(by=['month_N','Block'])
    final_df.drop(columns=['month_N'], inplace=True)

    # Формируем отчёт
    generate_report_with_charts(folder_path, final_df)

    return final_df





# Пример использования:
if __name__ == "__main__":
    tip = 1
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ЮЖР ИПГ 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Шатыркуль ИПГ 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ИПГ Саяк 3 2024'
    directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Жомарт ИПГ 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ИПГ Жайсан 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ВЖР ИПГ 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Жиланды ИПГ 2024'
    # directory =r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ЗР ИПГ 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Конырат ИПГ 2024'
    # directory = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__Акбастау ИПГ 2024"
    # directory = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__ИПГ 2024 С-1"
    # directory,tip = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__Нурказган ИПГ 2024",2   # ----N
    # directory,tip = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\Хаджиконган ИПГ 2024", 2   # ----N
    # directory = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__Абыз ИПГ 2024"

    final_data = load_excel_data_with_flex(directory, tip, max_rows=200)
    # print(final_data.head())
