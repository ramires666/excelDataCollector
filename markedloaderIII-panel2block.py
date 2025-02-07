import os
from datetime import datetime
import pandas as pd
from glob import glob
import re
import hashlib
import warnings
import numpy as np

import openpyxl
from openpyxl.utils import get_column_letter

from pymorphy2 import MorphAnalyzer

COL_LIMIT = 94

warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)


try:
    import win32com.client as win32
except ImportError:
    win32 = None  # On non-Windows or if Excel not installed

# Moth dictionary
morph = MorphAnalyzer()
months = {
    'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5, 'июнь': 6,
    'июль': 7, 'август': 8, 'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12
}


###############################################################################
#                         Conversion Functions
###############################################################################

def convert_xls_to_xlsx_with_formatting(xls_file_path, xlsx_file_path):
    """
    Converts a .xls file to .xlsx, preserving formatting (requires Windows + Excel).
    """
    if win32 is None:
        print("win32com not available; cannot convert .xls with formatting.")
        return
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(xls_file_path)
        workbook.SaveAs(xlsx_file_path, FileFormat=51)  # 51 -> .xlsx
        workbook.Close()
    except Exception as e:
        print(f"Ошибка при конвертации файла {xls_file_path}: {e}")
    finally:
        excel.Quit()


def convert_xlsb_to_xlsx_with_formatting(xlsb_file_path, xlsx_file_path):
    """
    Converts a .xlsb file to .xlsx, preserving formatting (requires Windows + Excel).
    """
    if win32 is None:
        print("win32com not available; cannot convert .xlsb with formatting.")
        return
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(xlsb_file_path)
        workbook.SaveAs(xlsx_file_path, FileFormat=51)  # 51 -> .xlsx
        workbook.Close()
    except Exception as e:
        print(f"Ошибка при конвертации файла {xlsb_file_path}: {e}")
    finally:
        excel.Quit()


###############################################################################
#                         Support / Helper Functions
###############################################################################

def generate_hash(s, length=2):
    """
    Генерирует хэш для строки s и возвращает первые length символов.
    """
    if not isinstance(s, str):
        s = str(s)
    h = hashlib.sha256(s.encode()).hexdigest()
    return h[:length]


def extract_month_from_filename(filename):
    """
    Extract Russian month from filename (using pymorphy2 for normal_form),
    returning (month_number, month_name).
    If not found -> (None, None).
    """
    basename = os.path.basename(filename)
    no_ext = os.path.splitext(basename)[0].lower()
    no_ext = re.sub(r'[^\w\s]', ' ', no_ext)
    no_ext = no_ext.replace('_', ' ')
    words = re.findall(r'\w+', no_ext)
    for w in words:
        parsed = morph.parse(w)[0]
        if parsed.normal_form in months:
            return months[parsed.normal_form], parsed.normal_form
    return None, None


def excel_column_to_index(column_letter):
    """
    Converts Excel column letter (A, B, C...) to zero-based index.
    """
    column_letter = column_letter.upper()
    index = 0
    for char in column_letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1


###############################################################################
#                    The "Yellow Column" and Related Logic
###############################################################################

import openpyxl
from openpyxl.cell import MergedCell

def get_yellow_columns_pandas_style(sheet):
    """
    Identifies columns in a given openpyxl sheet that contain >=10 cells
    with the exact fill color "FFFFFF00" (yellow), skipping hidden rows
    and merged cells.

    Returns a list of 0-based column indices. If we find any, we also insert
    the column before the first yellow at the front (like original logic).
    """
    max_col = sheet.max_column
    max_row = sheet.max_row

    col_color_count = [0]*max_col  # track # of yellow fills per column

    for row_idx in range(1, max_row+1):
        # skip hidden rows
        if sheet.row_dimensions[row_idx].hidden:
            continue

        # get all cells in this row
        row_cells = sheet[row_idx]
        # row_cells is e.g. a tuple of length = max_col

        for cell in row_cells:
            # skip merged cells (they're MergedCell, no col_idx)
            if isinstance(cell, MergedCell):
                continue

            # Now we can safely get the 1-based column index
            col_i = cell.col_idx - 1  # convert to 0-based
            if cell.fill and cell.fill.fgColor:
                fill_color = cell.fill.fgColor
                # Check if fill color is the "yellow" we expect
                if fill_color.type == 'rgb' and fill_color.rgb == 'FFFFFF00':
                    col_color_count[col_i] += 1

    # find which columns have >=10 yellow cells
    yellow_columns = [i for i, count in enumerate(col_color_count) if count >= 10]

    if yellow_columns:
        # Insert the column immediately before the first yellow
        first_yellow = yellow_columns[0]
        before_first = first_yellow - 1
        if before_first >= 0:
            yellow_columns.insert(0, before_first)

    return yellow_columns


def fill_merged_cells_in_first_yellow_column_in_memory(
        rows_data, merged_ranges, first_yellow_idx,
        row_to_data_index, max_cols
):
    """
    Fills merged cells in the first yellow column. If a range is strictly
    one column wide and multiple rows tall, we fill each row in that range
    with the top-left cell's value, skipping any rows not visible.
    """
    target_col = first_yellow_idx + 1  # 1-based in openpyxl

    for rng in merged_ranges:
        # Check if merges exactly one column
        if rng.min_col == target_col and rng.max_col == target_col:
            # We gather all the visible row indices
            top_left_row = rng.min_row
            top_left_val = None

            # The top-left row might be hidden or out of range
            if top_left_row in row_to_data_index:
                row_idx = row_to_data_index[top_left_row]
                # clamp columns
                if target_col <= max_cols:
                    top_left_val = rows_data[row_idx][target_col - 1]

            # fill sub-rows (min_row..max_row)
            for sheet_row in range(rng.min_row, rng.max_row + 1):
                if sheet_row in row_to_data_index:
                    row_idx = row_to_data_index[sheet_row]
                    if target_col <= max_cols:
                        rows_data[row_idx][target_col - 1] = top_left_val


def unmerge_horizontal_cells_in_memory(rows_data, merged_ranges, row_to_data_index, max_cols):
    """
    Unmerge horizontally in-memory if the top-left cell contains 'того'.
    row_to_data_index: dict mapping sheet row -> index in rows_data
    max_cols: maximum columns in each row (assuming you read entire row).
    """
    for rng in merged_ranges:
        # Check if horizontal (same row, multiple columns)
        if rng.min_row == rng.max_row and rng.min_col < rng.max_col:
            sheet_row = rng.min_row
            # Skip if that row is hidden or doesn't exist in rows_data
            if sheet_row not in row_to_data_index:
                continue

            row_idx = row_to_data_index[sheet_row]
            start_c = rng.min_col
            end_c   = rng.max_col

            # If the start col is beyond the data's columns, skip
            if start_c > max_cols:
                continue

            # Safely retrieve the top-left cell value (clamp it)
            if start_c <= max_cols:
                top_left_val = rows_data[row_idx][start_c - 1]
            else:
                top_left_val = None

            # If that cell contains "того", fill horizontally within range
            if top_left_val and ("того" in str(top_left_val).lower() or "сего" in str(top_left_val).lower()):
                for c in range(start_c, end_c + 1):
                    if c <= max_cols:
                        rows_data[row_idx][c - 1] = top_left_val



def forward_fill_column_by_index_in_memory(rows_data, col_idx):
    """
    Forward-fill the specified column in the rows_data 2D list
    (skipping any rows that are purely empty?).
    We'll treat the first non-empty encountered as the fill source
    for subsequent empty cells.
    """
    last_value = None
    for i in range(len(rows_data)):
        val = rows_data[i][col_idx]
        if val is not None and val != "":
            last_value = val
        else:
            # fill with last_value if it exists
            if last_value is not None:
                rows_data[i][col_idx] = last_value


def forward_fill_first_yellow_column_pandas(df, first_yellow_idx):
    """
    Forward-fills the 'first yellow' column in a DataFrame using Pandas .ffill().

    :param df: The DataFrame holding your sheet data (already created).
    :param first_yellow_idx: 0-based column index that is 'the first yellow column'
                             after you select columns in the DataFrame.
    """
    if df.empty:
        return df
    if first_yellow_idx < 0 or first_yellow_idx >= df.shape[1]:
        # out of range, do nothing
        return df

    col_name = df.columns[first_yellow_idx]
    df[col_name] = df[col_name].ffill()
    return df


###############################################################################
#                             Filtering Logic
###############################################################################

def is_valid_row(row):
    """
    Your old function checks if columns Ruda, Cu, Ag, fRuda, fCu, fAg
    are integer and not 0 -> skip or not.
    But you had 'return True' if at least one is not int or 0.
    We keep it so as you had it, or remove if not used.
    We'll keep it for reference, but you might prefer the new approach.
    """
    columns_to_check = ['Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg']
    for col in columns_to_check:
        val = row[col]
        if isinstance(val, int) and val != 0:
            # keep going
            continue
        # if at least one is not int or is 0, we consider "True" -> means we keep?
        return True
    return False


def parse_string(value):
    # Skip integer values
    if isinstance(value, int):
        return value

    # If the value isn't a string, convert it to a string
    if not isinstance(value, str):
        value = str(value)

    # Step 1: Strip leading and trailing whitespace
    value = value.strip()
    if not value:
        return value

    # Step 2 & 3: Keep the first character as is and convert the rest to lowercase.
    # This works for Russian letters as well.
    formatted_value = value[0] + value[1:].lower() if len(value) > 1 else value

    # Step 4: Insert a space between a digit and a letter following it.
    # The regex is updated to match both Latin and Cyrillic letters (including Ёё).
    formatted_value = re.sub(r'(?<=\d)(?=[A-Za-zА-Яа-яЁё])', ' ', formatted_value)

    return formatted_value


def fill_none_and_non_numeric_with_zero(df, columns_to_check=None):
    """
    Fills None values and non-numeric values with 0 in the specified columns.
    Default columns are 'Ruda' and 'fRuda'.
    """
    if columns_to_check is None:
        columns_to_check = ["Ruda", "fRuda"]

    # Create a copy of the DataFrame to avoid modifying the original
    filled_df = df.copy()

    # Fill None values with 0 in the specified columns
    filled_df[columns_to_check] = filled_df[columns_to_check].fillna(0)

    # Replace non-numeric values with 0 in the specified columns
    for col in columns_to_check:
        filled_df[col] = pd.to_numeric(filled_df[col], errors='coerce').fillna(0)

    return filled_df
# def fill_none_and_filter_non_numeric(df, columns_to_check=None):
#     """
#     Fills None values with 0 and filters out rows with non-numeric values in the specified columns.
#     Default columns are 'Ruda' and 'fRuda'.
#     """
#     if columns_to_check is None:
#         columns_to_check = ["Ruda", "fRuda"]
#
#     # Create a copy of the DataFrame to avoid modifying the original
#     filled_df = df.copy()
#
#     # Fill None values with 0 in the specified columns
#     filled_df[columns_to_check] = filled_df[columns_to_check].fillna(0)
#
#     # Filter out rows with non-numeric values in the specified columns
#     mask = (
#         filled_df[columns_to_check]
#         .apply(pd.to_numeric, errors='coerce')  # Convert to numeric, non-numeric become NaN
#         .isna()  # Check for NaN (non-numeric values)
#         .any(axis=1)  # Keep rows where any column has NaN
#     )
#
#     # Keep rows without non-numeric values
#     filtered_df = filled_df[~mask]
#
#     return filtered_df


def filter_columns_with_whole_number_sums(df, columns_to_check=None):
    """
    Removes rows where the sum of the specified columns is an integer
    (unchanged logic), but does not remove rows where the sum is greater than 150.
    """
    if columns_to_check is None:
        columns_to_check = ["Ruda", "Cu", "Ag", "fRuda", "fCu", "fAg"]

    filtered_df = df.copy()

    for index, row in filtered_df.iterrows():
        row_sum = 0
        has_non_numeric = False

        for col in columns_to_check:
            value = row[col]

            # Treat None as 0
            if pd.isna(value) or value == '' or value is None:
                value = 0

            try:
                row_sum += float(value)
            except (ValueError, TypeError):
                has_non_numeric = True
                break

        # Drop rows where the sum is an integer (including 0) and the sum is not greater than 150
        if (not has_non_numeric) and (row_sum == int(row_sum)) and (row_sum <= 350):
            filtered_df.drop(index, inplace=True)

    return filtered_df


###############################################################################
#               Calculations & Aggregations (unchanged)
###############################################################################

def add_calculated_columns(df):
    """
    Adds difference columns, ratio columns, etc.
    """
    cols = ['Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg']
    for c in cols:
        df[c] = df[c].replace([None, ''], np.nan)
        df[c] = df[c].astype(str).str.replace(',', '').str.strip()
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    df['diff-Ruda'] = abs(df['Ruda'] - df['fRuda'])
    df['diff-Cu'] = abs(df['Cu'] - df['fCu'])
    df['diff-Ag'] = abs(df['Ag'] - df['fAg'])

    # 1-(diff/val) clipped [0..100]
    df['1-(Ruda-fRuda)/Ruda'] = np.clip(np.where(df['Ruda'] != 0, 1 - df['diff-Ruda'] / df['Ruda'], 1.0), 0, 100)
    df['1-(Cu-fCu)/Cu'] = np.clip(np.where(df['Cu'] != 0, 1 - df['diff-Cu'] / df['Cu'], 1.0), 0, 100)
    df['1-(Ag-fAg)/Ag'] = np.clip(np.where(df['Ag'] != 0, 1 - df['diff-Ag'] / df['Ag'], 1.0), 0, 100)

    # %Cu, %fCu
    df['%Cu'] = np.where(df['Ruda'] != 0, df['Cu'] / df['Ruda'] * 100, 0)
    df['%fCu'] = np.where(df['fRuda'] != 0, df['fCu'] / df['fRuda'] * 100, 0)

    # 1-(%Cu-%fCu)/%Cu
    df['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where(df['%Cu'] != 0, 1 - abs((df['%Cu'] - df['%fCu']) / df['%Cu']), 1.0),
        0, 100
    )

    # p>0, f>0
    df['p>0'] = np.where(df['Ruda'] > 0, 1, 0)
    df['f>0'] = np.where(df['fRuda'] > 0, 1, 0)

    df = filter_columns_with_whole_number_sums(df)

    return df


def group_by_block_and_month(df):
    sum_cols = ['Ruda', 'fRuda', 'Cu', 'fCu', 'Ag', 'fAg']
    grouped = df.groupby(['Panel', 'month'], as_index=False)[sum_cols].sum()

    grouped['Ruda-fRuda'] = grouped['Ruda'] - grouped['fRuda']
    grouped['Cu-fCu'] = grouped['Cu'] - grouped['fCu']
    grouped['Ag-fAg'] = grouped['Ag'] - grouped['fAg']

    # ratio columns
    grouped['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(grouped['Ruda'] != 0, 1 - abs(grouped['Ruda-fRuda'] / grouped['Ruda']), 1.0), 0, 100
    )
    grouped['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(grouped['Cu'] != 0, 1 - abs(grouped['Cu-fCu'] / grouped['Cu']), 1.0), 0, 100
    )
    grouped['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(grouped['Ag'] != 0, 1 - abs(grouped['Ag-fAg'] / grouped['Ag']), 1.0), 0, 100
    )

    grouped['%Cu'] = np.where(grouped['Ruda'] != 0, grouped['Cu'] / grouped['Ruda'] * 100, 0)
    grouped['%fCu'] = np.where(grouped['fRuda'] != 0, grouped['fCu'] / grouped['fRuda'] * 100, 0)

    grouped['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where(grouped['%Cu'] != 0, 1 - abs((grouped['%Cu'] - grouped['%fCu']) / grouped['%Cu']), 1.0),
        0, 100
    )

    grouped['p>0'] = np.where(grouped['Ruda'] > 0, 1, 0)
    grouped['f>0'] = np.where(grouped['fRuda'] > 0, 1, 0)

    grouped = filter_columns_with_whole_number_sums(grouped)


    return grouped


def group_by_block(df):
    sum_cols = ['Ruda', 'fRuda', 'Cu', 'fCu', 'Ag', 'fAg']
    grouped = df.groupby(['Panel'], as_index=False)[sum_cols].sum()

    grouped['Ruda-fRuda'] = grouped['Ruda'] - grouped['fRuda']
    grouped['Cu-fCu'] = grouped['Cu'] - grouped['fCu']
    grouped['Ag-fAg'] = grouped['Ag'] - grouped['fAg']

    grouped['%Cu'] = np.where(grouped['Ruda'] != 0, grouped['Cu'] / grouped['Ruda'] * 100, 0)
    grouped['%fCu'] = np.where(grouped['fRuda'] != 0, grouped['fCu'] / grouped['fRuda'] * 100, 0)

    grouped['1-(Ruda-fRuda)/Ruda'] = np.clip(
        np.where(grouped['Ruda'] != 0, 1 - abs(grouped['Ruda-fRuda'] / grouped['Ruda']), 1.0), 0, 100
    )
    grouped['1-(Cu-fCu)/Cu'] = np.clip(
        np.where(grouped['Cu'] != 0, 1 - abs(grouped['Cu-fCu'] / grouped['Cu']), 1.0), 0, 100
    )
    grouped['1-(Ag-fAg)/Ag'] = np.clip(
        np.where(grouped['Ag'] != 0, 1 - abs(grouped['Ag-fAg'] / grouped['Ag']), 1.0), 0, 100
    )
    grouped['1-(%Cu-%fCu)/%Cu'] = np.clip(
        np.where(grouped['%Cu'] != 0, 1 - abs((grouped['%Cu'] - grouped['%fCu']) / grouped['%Cu']), 1.0),
        0, 100
    )

    grouped['p>0'] = np.where(grouped['Ruda'] > 0, 1, 0)
    grouped['f>0'] = np.where(grouped['fRuda'] > 0, 1, 0)

    grouped = filter_columns_with_whole_number_sums(grouped)

    return grouped


def mean_values_groupped_by_month(df):
    """
    Groups by month, taking mean of ratio columns.
    """
    needed = ['1-(Ruda-fRuda)/Ruda', '1-(Cu-fCu)/Cu', '1-(Ag-fAg)/Ag', '1-(%Cu-%fCu)/%Cu']
    grouped = df.groupby('month', as_index=False)[needed].mean()
    return grouped


# def assign_block_numbers(data, col_idx):
#     """
#     Assigns consecutive block numbers from top to bottom:
#       - Every time we see a row containing 'того', that row is a delimiter
#         (i.e. not assigned any block).
#       - All preceding non-'того' rows belong to the current block number.
#       - After a 'того' row, we increment the block number for subsequent rows.
#
#     Returns a modified DataFrame with a 'Block' column and
#     rows containing 'того' removed.
#     """
#     # Convert the relevant column to string, check for 'того'
#     is_togo = data.iloc[:, col_idx].astype(str).str.lower().str.contains('того', na=False)
#
#     block_assignments = []
#     block_num = 1
#
#     for has_togo in is_togo:
#         if has_togo:
#             # This row is a delimiter => no block
#             block_assignments.append(None)
#             # Increment block counter for the next row
#             block_num += 1
#         else:
#             # This row belongs to current block
#             block_assignments.append(block_num)
#
#     data['Block'] = block_assignments
#
#     # Finally, remove the 'того' rows
#     data = data[data['Block'].notna()]
#
#     # Convert the block column to integer
#     data['Block'] = data['Block'].astype(int)
#
#     return data


import pandas as pd

def assign_block_numbers_fcking_LOOONG(data, col_idx):
    """
    Assigns consecutive block numbers from top to bottom:
      - Every time we see a row containing 'того', that row is a delimiter
        (i.e., not assigned any block).
      - The block number is incremented only if:
          1. The current row contains 'того'.
          2. There is at least one non-'того' row before the current 'того' row.
      - All preceding non-'того' rows belong to the current block number.

    Returns a modified DataFrame with a 'Block' column,
    and rows containing 'того' removed.
    """
    # Convert the relevant column to string, check for 'того'
    is_togo = data.iloc[:, col_idx].astype(str).str.lower().str.contains('того', na=False)

    block_assignments = []
    block_num = 1
    has_non_togo_before = False

    for has_togo in is_togo:
        if has_togo:
            # This row is a delimiter => no block
            block_assignments.append(None)
            # Increment block counter only if there is at least one non-'того' row before
            if has_non_togo_before:
                block_num += 1
                has_non_togo_before = False
        else:
            # This row belongs to the current block
            block_assignments.append(block_num)
            has_non_togo_before = True

    # Assign the block list
    data['Block'] = block_assignments

    # Remove the 'того' rows => create a real copy
    data = data[data['Block'].notna()].copy()

    # Convert 'Block' to integer
    data['Block'] = data['Block'].astype(int)

    return data


def assign_block_numbers(data, col_idx):
    """
    Для DataFrame data с уже выбранными "жёлтыми" столбцами.
    col_idx – индекс столбца (обычно col_before_yellow), по которому определяется,
    содержит ли строка слово "того" (делимитер).

    Функция перебирает строки, назначая каждой не-делимитной строке имя панели,
    которое берётся из первого столбца (индекс 0) первой строки блока.
    Строки с делимитером удаляются.
    """
    # Определяем, является ли строка делимитером (содержит слово "того")
    # data = filter_columns_with_whole_number_sums(data)

    is_delimiter = data.iloc[:, col_idx].astype(str).str.lower().str.contains('того', na=False)

    panel_names = []
    current_panel = None

    # Перебираем строки DataFrame
    for idx, is_delim in enumerate(is_delimiter):
        if is_delim:
            # Если строка является делимитером, сбрасываем текущую панель
            panel_names.append(None)
            current_panel = None
        else:
            # Если текущая панель не установлена, устанавливаем её из первого столбца этой строки
            if current_panel is None:
                current_panel = data.iloc[idx, 1]
            panel_names.append(current_panel)

    # Создаем копию DataFrame и добавляем столбец с именами панели
    data = data.copy()
    data['Block'] = panel_names

    # Удаляем строки с делимитером (где имя панели не назначено)
    data = data.loc[data['Block'].notna()].copy()

    return data


###############################################################################
#                   Master Function to Load with Minimal Openpyxl
###############################################################################

def load_excel_data_with_flex(folder_path, tip=1, max_rows=500):
    """
    1) Convert all .xls/.xlsb -> .xlsx if needed.
    2) For each .xlsx:
       - Use openpyxl to get hidden rows, merges, color-based columns, etc.
       - Build a 2D list skipping hidden rows, unmerging in memory, etc.
       - Identify the "yellow columns", forward-fill the first one, etc.
       - Convert to Pandas, rename columns, skip "того",
         filter out integer-sum rows, add 'Block' cumsum, etc.
    3) Combine into final_df, do aggregator, output a single final .xlsx.
    """
    # Step 1: Convert
    xls_files = glob(os.path.join(folder_path, "*.xls"))
    for xls in xls_files:
        out_xlsx = os.path.splitext(xls)[0] + ".xlsx"
        if not os.path.exists(out_xlsx):
            convert_xls_to_xlsx_with_formatting(xls, out_xlsx)

    xlsb_files = glob(os.path.join(folder_path, "*.xlsb"))
    for xlsb in xlsb_files:
        out_xlsx = os.path.splitext(xlsb)[0] + ".xlsx"
        if not os.path.exists(out_xlsx):
            convert_xlsb_to_xlsx_with_formatting(xlsb, out_xlsx)

    # Step 2: Now read .xlsx
    xlsx_files = glob(os.path.join(folder_path, "*.xlsx"))
    final_columns = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg', 'Uchastok', 'month',
                     'Block']
    final_df = pd.DataFrame(columns=final_columns)

    for file_path in xlsx_files:
        file_name = os.path.basename(file_path)
        month_num, month_name = extract_month_from_filename(file_name)
        if not month_num:
            print(f"-> -> ->  нет месяца - пропускаю файл: {file_name}")
            continue

        # open workbook in read-only or normal mode
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        except:
            print(f"Cannot open {file_path}")
            continue

        # skip hidden sheets
        visible_sheets = [sh for sh in wb.sheetnames if wb[sh].sheet_state == "visible"]

        # if > 1 visible sheet, optionally skip the first. The old code did that:
        if len(visible_sheets) > 1:
            visible_sheets = visible_sheets[1:]  # just replicate your logic

        for sheet_name in visible_sheets:
            ws = wb[sheet_name]
            if ws.max_row == 0 or ws.max_column == 0:
                print(f"Sheet '{sheet_name}' is empty in {file_name}")
                continue

            print(f"\n>>>>>>>>>>>>>>>> {file_name} => sheet '{sheet_name}' => month {month_name}")



            # Enforce row limit
            max_row_to_process = min(ws.max_row, max_rows)

            # 2.1) Identify hidden rows, merges, yellow columns
            # merges
            merged_ranges = list(ws.merged_cells.ranges)

            # yellow columns (0-based)
            yellow_cols = get_yellow_columns_pandas_style(ws)
            if not yellow_cols:
                print(f"No 'yellow' columns found in '{sheet_name}' => skipping")
                continue

            first_yellow_idx = yellow_cols[0]  # 0-based

            # 2.2) Build rows_data & row_to_data_index (skip hidden rows)
            visible_rows = []
            rows_data = []
            # for row_num in range(1, ws.max_row+1):
            for row_num in range(1, max_row_to_process  + 1):
                if ws.row_dimensions[row_num].hidden:
                    continue  # skip hidden
                row_cells = ws[row_num]
                row_vals = []
                for cell in row_cells:
                    if isinstance(cell, MergedCell):
                        row_vals.append(None)
                    else:
                        row_vals.append(cell.value)
                rows_data.append(row_vals)
                visible_rows.append(row_num)

            if not rows_data:
                print(f"After skipping hidden rows, no data in {sheet_name}.")
                continue

            row_to_data_index = { real_row: i for i, real_row in enumerate(visible_rows) }
            max_col_count = max(len(r) for r in rows_data)

            try:
                # unmerge horizontally
                unmerge_horizontal_cells_in_memory(rows_data, merged_ranges, row_to_data_index, max_col_count+1)
                "yep"
            except:
                unmerge_horizontal_cells_in_memory(rows_data, merged_ranges, row_to_data_index, max_col_count)

            # fill merged cells in the first 3 yellow columns (like your old code)
            # fill_merged_cells_in_first_yellow_column_in_memory(
            #     rows_data, merged_ranges, first_yellow_idx,
            #     row_to_data_index, max_col_count
            # )
            # if len(yellow_cols) > 1:
            #     fill_merged_cells_in_first_yellow_column_in_memory(
            #         rows_data, merged_ranges, first_yellow_idx+1,
            #         row_to_data_index, max_col_count
            #     )
            # if len(yellow_cols) > 2:
            #     fill_merged_cells_in_first_yellow_column_in_memory(
            #         rows_data, merged_ranges, first_yellow_idx+2,
            #         row_to_data_index, max_col_count
            #     )
            # forward fill the first yellow column
            # forward_fill_column_by_index_in_memory(rows_data, first_yellow_idx)

            # 2.3) Convert 2D list to Pandas
            # The first row is presumably headers:
            header_row = [str(x) if x else "" for x in rows_data[0]]
            data_body = rows_data[1:]  # skip header
            if not data_body:
                print(f"No data rows in sheet '{sheet_name}' after removing header row.")
                continue

            data = pd.DataFrame(data_body, columns=header_row)

            data = data.dropna(how='all')

            # use only columns at the "yellow_cols" indices
            # (the code: data = data.iloc[:, yellow_cols]), but we must clamp if out of range
            max_index = data.shape[1] - 1
            valid_yellows = [c for c in yellow_cols if 0 <= c <= max_index]
            if not valid_yellows:
                print(f"yellow cols out of range in sheet {sheet_name}? skipping")
                continue
            data = data.iloc[:, valid_yellows]


            data.iloc[:, 0] = data.iloc[:, 0].ffill()

            data.iloc[:, 1] = data.iloc[:, 1].ffill()
            data.iloc[:, 2] = data.iloc[:, 2].ffill()

            # data = forward_fill_first_yellow_column_pandas(data, 0)

            # old code: col_before_yellow = yellow_cols[0] - 1 => used to parse 'ИТОГО', etc.
            # but we have that logic in unmerge. We'll do the same:
            col_before_yellow_idx = valid_yellows[0] - 1
            if col_before_yellow_idx < 0:
                col_before_yellow_idx = 0

            # # Identify rows containing "того"
            is_togo = data.iloc[:, col_before_yellow_idx].astype(str).str.lower().str.contains('того', na=False)
            is_vsego = data.iloc[:, col_before_yellow_idx].astype(str).str.lower().str.contains('сего', na=False)
            #
            # # Build consecutive blocks
            # block_list = []
            # block_counter = 1
            # for flag in is_togo:
            #     block_list.append(block_counter)
            #     if flag:
            #         block_counter += 1

            # data['Block'] = block_list
            data['Block'] = None


            # set columns based on OGR or tip
            if 'ОГР' in sheet_name.upper():
                std_cols = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'fRuda', 'fCu', 'Ag', 'fAg', 'Block']
                data.columns = std_cols

                # ogr_cols = ['Shtrek','Panel', 'Ruda', 'Cu', 'fRuda', 'fCu', 'Ag', 'fAg']
                # if len(data.columns) == len(ogr_cols):
                #     data.columns = ogr_cols
                #     # data['Shtrek'] = ''
                #     data['Horizont'] = ''
                # else:
                #     print(f"Warning: OGR sheet {sheet_name} => unexpected # of columns.")
                #     continue
            else:
                if tip == 1:
                    std_cols = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'Ag', 'fRuda', 'fCu', 'fAg', 'Block']
                else:
                    # tip=2
                    std_cols = ['Horizont', 'Panel', 'Shtrek', 'Ruda', 'Cu', 'fRuda', 'fCu', 'Ag', 'fAg', 'Block']

                if len(data.columns) == len(std_cols):
                    data.columns = std_cols
                else:
                    print(f"Warning: sheet '{sheet_name}' => unexpected # of columns.")
                    continue

            # remove the first row (like original code does .iloc[1:])
            data = data.iloc[1:] if len(data) > 1 else data.iloc[0:0]




            data = fill_none_and_non_numeric_with_zero(data)



            # skip empty shtrek or panel
            if 'Shtrek' in data.columns:
                data = data[data['Shtrek'].notnull() & (data['Shtrek'] != '')]
            if 'Panel' in data.columns:
                data = data[data['Panel'].notnull() & (data['Panel'] != '')]

            # add 'Uchastok', 'month'
            data['Uchastok'] = sheet_name
            data['month'] = month_name

            # ensure all final columns exist
            for col in final_columns:
                if col not in data.columns:
                    data[col] = ''



            # remove rows containing 'того' in Panel/Shtrek/Horizont
            for c in ['Panel', 'Shtrek', 'Horizont']:
                if c in data.columns:
                    data = data[~data[c].astype(str).str.lower().str.contains('того', na=False)]
                    data = data[~data[c].astype(str).str.lower().str.contains('сего', na=False)]

            # remove rows if Ruda=0 & fRuda=0
            data = data[~((data['Ruda'] == 0) & (data['fRuda'] == 0))]
            # # filter columns with whole number sums
            data = filter_columns_with_whole_number_sums(data)
            # remove rows if fAg == 'кг'?
            data = data[~(data['fAg'] == 'кг')]

            data['Panel'] = data['Panel'].apply(parse_string)
            # data['Block'] = data['Panel'].str.strip()
            # data = assign_block_numbers(data, col_before_yellow_idx)
            # Remove the rows that contain 'того'
            data = data[~is_togo]
            data = data[~is_vsego]
            data = filter_columns_with_whole_number_sums(data)



            # reorder columns
            data = data[final_columns]
            final_df = pd.concat([final_df, data], ignore_index=True)

    # after collecting
    final_df = final_df[~((final_df['Ruda'] == 0) & (final_df['fRuda'] == 0))]

    # # filter columns with whole number sums
    final_df = filter_columns_with_whole_number_sums(final_df)

    # order by month
    final_df['month_N'] = final_df['month'].map(months)
    final_df = final_df.sort_values(by=['month_N', 'Block'])
    final_df.drop(columns=['month_N'], inplace=True)

    # do aggregator & produce final result
    generate_report_with_charts(folder_path, final_df)
    return final_df


###############################################################################
#       Final Reporting (modified to rely on Pandas or minimal openpyxl)
###############################################################################

def generate_report_with_charts(folder_path, full_df):
    """
    The final aggregator steps,
    then write to a single xlsx with charts (using xlsxwriter).
    """
    full_df = add_calculated_columns(full_df)

    block_and_month_aggregated_df = group_by_block_and_month(full_df)
    block_aggregated_df = group_by_block(full_df)
    monthly_avg_df = mean_values_groupped_by_month(block_and_month_aggregated_df)

    # naming
    t = datetime.now().microsecond
    output_file_name = f"_report_{os.path.basename(folder_path.strip('/\\'))}{t}.xlsx"
    output_file = os.path.join(folder_path, output_file_name)

    # Write with xlsxwriter
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        full_df.to_excel(writer, sheet_name="По штрекам", index=False)
        block_and_month_aggregated_df.to_excel(writer, sheet_name="Сумм. панель-месяц", index=False)
        block_aggregated_df.to_excel(writer, sheet_name="Сумм. по панель", index=False)
        monthly_avg_df.to_excel(writer, sheet_name="Средн. по месяцам", index=False)

        # optionally create a chart sheet or simply rely on these raw sheets
        workbook = writer.book
        worksheet = workbook.add_worksheet("Some Charts")
        # Example chart creation if you want

    print(f"\n==> Итоговый отчет сохранен: {output_file}")

# Example usage at bottom if you want:
if __name__ == "__main__":
    tip = 1
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ЮЖР ИПГ 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Шатыркуль ИПГ 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ИПГ Саяк 3 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Жомарт ИПГ 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ИПГ Жайсан 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ВЖР ИПГ 2024'
    directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Жиланды ИПГ 2024'
    # directory =r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__ЗР ИПГ 2024'
    # directory = r'C:\Users\delxps\Documents\Kazakhmys\_alibek\__Конырат ИПГ 2024'
    # directory = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__Акбастау ИПГ 2024"
    # directory = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__ИПГ 2024 С-1"
    # directory,tip = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__Нурказган ИПГ 2024",2   # ----N
    # directory,tip = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\Хаджиконган ИПГ 2024",2   # ----N
    # directory = r"C:\Users\delxps\Documents\Kazakhmys\_alibek\__Абыз ИПГ 2024"

    final_data = load_excel_data_with_flex(directory, tip, max_rows=200)
    # print(final_data.head())
