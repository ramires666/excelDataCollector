import zipfile
import pandas as pd
import os
import glob
from openpyxl.utils import get_column_letter  # Correct import for get_column_letter

# ----------------------------
# Global Configuration Variables
# ----------------------------

# Define the bus prefix (change this to 'gruz' or any other name as needed)
BUS_PREFIX = 'spec'  # Change to 'bus', 'gruz', etc.

# Define the directory containing the ZIP archives
ARCHIVES_DIR = r'C:\Users\delxps\PycharmProjects\excelCollector'  # Update if necessary

# Define the output Excel file path, including BUS_PREFIX
OUTPUT_EXCEL = os.path.join(ARCHIVES_DIR, f'{BUS_PREFIX}_aggregated_sliv_data.xlsx')

# Mapping of month abbreviations to their numeric representations
MONTH_ABBR_TO_NUM = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}

# List of columns that contain liter measurements and need cleaning
LITERS_COLUMNS = ['Нач. уровень', 'Слито', 'Кон. уровень']

# List of other numeric columns to be converted
OTHER_NUMERIC_COLUMNS = ['Кол-во', 'Счетчик']

# Define the bus column name in the CSV (change if different)
BUS_COLUMN_NAME = 'Группировка'  # Change if your CSV uses a different column name

# ----------------------------
# Function Definitions
# ----------------------------

def extract_month_abbr(zip_filename):
    """
    Extracts the month abbreviation from the ZIP filename.
    Assumes the filename format is '<BUS_PREFIX>_sliv_<mon>.zip'

    :param zip_filename: Name of the ZIP file
    :return: Month abbreviation as a string
    """
    basename = os.path.basename(zip_filename)
    try:
        parts = basename.split('_')
        if len(parts) != 3:
            raise ValueError(f"Unexpected filename format: {basename}")
        month_abbr_with_ext = parts[2]
        month_abbr = os.path.splitext(month_abbr_with_ext)[0].lower()
        return month_abbr
    except Exception as e:
        print(f"Error extracting month from '{basename}': {e}")
        return None

def read_csv_from_zip(zip_path, csv_filename):
    """
    Reads a CSV file from a ZIP archive into a pandas DataFrame.

    :param zip_path: Path to the ZIP file
    :param csv_filename: Name of the CSV file inside the ZIP
    :return: pandas DataFrame or None if failed
    """
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            if csv_filename in zip_ref.namelist():
                with zip_ref.open(csv_filename) as csvfile:
                    # No encoding specified as per requirement
                    df = pd.read_csv(csvfile, delimiter=';')
                    return df
            else:
                print(f"CSV file '{csv_filename}' not found in '{zip_path}'.")
                return None
    except zipfile.BadZipFile:
        print(f"Error: '{zip_path}' is not a valid ZIP file.")
        return None
    except Exception as e:
        print(f"Error reading '{csv_filename}' from '{zip_path}': {e}")
        return None

def clean_liters_columns(df, columns):
    """
    Removes " л" from specified liters columns and converts them to numeric types.

    :param df: pandas DataFrame
    :param columns: List of column names to clean
    :return: Cleaned DataFrame
    """
    for col in columns:
        if col in df.columns:
            # Remove " л" and any surrounding whitespace
            df[col] = df[col].astype(str).str.replace(' л', '', regex=False).str.strip()
            # Replace '-----' or other non-numeric placeholders with NaN
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df

def clean_other_numeric_columns(df, columns):
    """
    Converts specified columns to numeric types, handling any non-numeric values.

    :param df: pandas DataFrame
    :param columns: List of column names to convert
    :return: Cleaned DataFrame
    """
    for col in columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df

def add_month_columns(df, month_abbr):
    """
    Adds 'Month_Abbr' and 'Month_Num' columns to the DataFrame.

    :param df: pandas DataFrame
    :param month_abbr: Three-letter month abbreviation
    :return: DataFrame with added columns
    """
    df['Month_Abbr'] = month_abbr.capitalize()  # e.g., 'Nov'
    month_num = MONTH_ABBR_TO_NUM.get(month_abbr.lower(), None)
    df['Month_Num'] = month_num
    return df

def main():
    # Pattern to match ZIP files like bus_sliv_nov.zip or gruz_sliv_nov.zip
    zip_pattern = os.path.join(ARCHIVES_DIR, f'{BUS_PREFIX}_sliv_*.zip')
    zip_files = glob.glob(zip_pattern)

    # Debug: List all found ZIP files
    print(f"Found {len(zip_files)} ZIP file(s) matching the pattern '{BUS_PREFIX}_sliv_*.zip':")
    for file in zip_files:
        print(f" - {file}")

    if not zip_files:
        print(f"No ZIP archives found matching the pattern '{BUS_PREFIX}_sliv_*.zip' in '{ARCHIVES_DIR}'.")
        return

    aggregated_data = []  # List to hold all DataFrames

    for zip_path in zip_files:
        print(f"\nProcessing archive: {zip_path}")
        month_abbr = extract_month_abbr(zip_path)

        if not month_abbr:
            print(f"Skipping '{zip_path}' due to extraction error.")
            continue

        # Validate month abbreviation
        if month_abbr.lower() not in MONTH_ABBR_TO_NUM:
            print(f"Unknown month abbreviation '{month_abbr}' in '{zip_path}'. Skipping.")
            continue

        # Construct the expected CSV filename
        csv_filename = f"{BUS_PREFIX}_sliv_{month_abbr}_Сливы.csv"

        df = read_csv_from_zip(zip_path, csv_filename)

        if df is not None:
            # Clean liters columns by removing " л" and converting to numbers
            df = clean_liters_columns(df, LITERS_COLUMNS)

            # Clean other numeric columns
            df = clean_other_numeric_columns(df, OTHER_NUMERIC_COLUMNS)

            # Add 'Month_Abbr' and 'Month_Num' columns
            df = add_month_columns(df, month_abbr)

            # Check if BUS_COLUMN_NAME exists, else fill with 'Unknown'
            if BUS_COLUMN_NAME not in df.columns:
                print(f"Warning: '{BUS_COLUMN_NAME}' column not found in '{csv_filename}'. Filling with 'Unknown'.")
                df[BUS_COLUMN_NAME] = 'Unknown'

            # Append to the list
            aggregated_data.append(df)
            print(f"Successfully processed '{csv_filename}'. Rows added: {len(df)}")
        else:
            print(f"Failed to process '{csv_filename}' from '{zip_path}'.")

    if aggregated_data:
        # Concatenate all DataFrames
        combined_df = pd.concat(aggregated_data, ignore_index=True)

        # Optional: Replace '-----' with NaN or other placeholder if any remain
        combined_df.replace('-----', pd.NA, inplace=True)

        # Reorder columns to place 'Month_Num' and 'Month_Abbr' first, then bus column, then the rest
        if BUS_COLUMN_NAME in combined_df.columns:
            cols = ['Month_Num', 'Month_Abbr', BUS_COLUMN_NAME] + [col for col in combined_df.columns if col not in ['Month_Num', 'Month_Abbr', BUS_COLUMN_NAME]]
        else:
            cols = ['Month_Num', 'Month_Abbr'] + [col for col in combined_df.columns if col not in ['Month_Num', 'Month_Abbr']]
        combined_df = combined_df[cols]

        # Save to Excel with appropriate data types
        try:
            # Use ExcelWriter to specify the engine and ensure proper data types
            with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
                combined_df.to_excel(writer, index=False)

                # Access the workbook and worksheet to set number formats if needed
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # Example: Set number formats for specific columns
                for idx, column in enumerate(combined_df.columns, 1):
                    if column in LITERS_COLUMNS + OTHER_NUMERIC_COLUMNS + ['Month_Num']:
                        # Get Excel column letter
                        col_letter = get_column_letter(idx)
                        # Apply number format with two decimal places
                        # Note: This sets the entire column; adjust as needed
                        worksheet.column_dimensions[col_letter].number_format = '0.00'

            print(f"\nAggregated data saved successfully to '{OUTPUT_EXCEL}'.")
        except Exception as e:
            print(f"Error saving to Excel: {e}")
    else:
        print("No data was aggregated. Please check the ZIP archives and CSV files.")

if __name__ == "__main__":
    main()
